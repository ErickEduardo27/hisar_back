from django.shortcuts import render
from django.shortcuts import get_object_or_404
from Asistencial.models import incidenciaEnfermeriaDetalle,procedimientoEnfermeria,incidenciaEnfermeriaCabecera,programacionTurno,unidadAtencion,correos,solicitud,movimientoPaciente,pacienteLocalizacion,pacienteGeoTem,ubigeo,asisPacDiarioAdicional,examenLaboratorio,dpDiario,loginAppHisar,serologiaPaciente,baseDatosProduccion,asisPacDiario,asigCuposPac,parameCentroPuesto,parameCentro,docuContratados,listaEspera,parNuticion,maestroMatSap,delegacionBienesEstra, cas, usuario,perfil, paciente, examen, archivo, bienAmbiente, bienImag, presAnemia,admiAnemia, exclusionAnemia, movimientoAnemia, bienPersonal, bienpat, dependencia, ambiente, personal, proveedor, provMaq, incidenciaDsi, maestro, bienHadware, bienSoftware, bienDetalleMonitor, nutricion, personalVpn, personalCertificado, valGlobalSub,formularioCambioClinica,hospital,medico,formularioCapacitacion,laboratorio,instaladores,protocoloAnemia,protocoloTmo,protocoloNutricion,laboratorioTemp,formularioCenso,registroCarga,laboratorioKtv
from Asistencial.serializers import incidenciaEnfermeriaDetalleSerializer,procedimientoEnfermeriaSerializer,incidenciaEnfermeriaCabeceraSerializer,programacionTurnoSerializer,unidadAtencionSerializer,correosSerializer,solicitudSerializer,movimientoPacienteSerializer,pacienteLocalizacionSerializer,pacienteGeoTemSerializer,ubigeoSerializer,asisPacDiarioAdicionalSerializer,examenLaboratorioSerializer,dpDiarioSerializer,loginAppHisarSerializer,serologiaPacienteSerializer,baseDatosProduccionSerializer,asisPacDiarioSerializer,asigCuposPacSerializer,parameCentroPuestoSerializer,parameCentroSerializer,docuContratadosSerializer,listaEsperaSerializer,parNuticionSerializer,maestroMatSapSerializer,delegacionBienesEstraSerializer, casSerializer ,usuarioSerializer,perfilSerializer, PacienteSerializer, ExamenSerializer, ArchivoSerializer, presAnemiaSerializer, admiAnemiaSerializer, exclusionAnemiaSerializer, movimientoAnemiaSerializer, bienAmbienteSerializer, bienImagSerializer, bienPersonalSerializer, bienpatSerializer, dependenciaSerializer, ambienteSerializer, personalSerializer, proveedorSerializer, provMaqSerializer, incidenciaDsiSerializer, maestroSerializer, bienHadwareSerializer, bienSoftwareSerializer, bienDetalleMonitorSerializer, nutricionSerializer, personalVpnSerializer, personalCertificadoSerializer, valGlobalSubSerializer,formularioCambioClinicaSerializer,hospitalSerializer,medicoSerializer,formularioCapacitacionSerializer,laboratorioSerializer,instaladoresSerializer,protocoloAnemiaSerializer,protocoloTmoSerializer,protocoloNutricionSerializer,laboratorioTempSerializer,registroCargaSerializer
from rest_framework import permissions, viewsets, filters
from django.db import transaction
from bs4 import BeautifulSoup
import re
from datetime import datetime
from rest_framework.pagination import PageNumberPagination
from django.db.models import Q
#servicios externos
from django.http import HttpResponse,JsonResponse
import requests
import json
from decouple import config
# Create your views here.
from rest_framework.decorators import api_view
from django.contrib.auth.decorators import permission_required
from rest_framework.permissions import IsAuthenticated

from django.core.mail import EmailMessage
from django.conf import settings
from datetime import date
from django.db import connection
import openpyxl
import xlrd
import json
import os
import socket
import time
from decimal import Decimal

@api_view(['POST'])
@permission_required([IsAuthenticated])
def rep_Informe_Nutricion_Admin(request):
    body=request.body.decode('utf-8')
    objeto_python = json.loads(body)
    #Calculamos la fecha de hoy
    hoy = date.today()
    fecha_formateada = hoy.strftime("%m-%d-%Y")
    if objeto_python["fecha"]==None :
        fecha=fecha_formateada
    else :
        fecha=objeto_python["fecha"]

    # Importar las clases necesarias
    cursor = connection.cursor()
    sql = 'select "descripCas",usu."nombre",to_char("fechaReg",'+"'YYYY-MM-DD'"+') ,concat("ape_pat",'+"' '"+',"ape_mat",'+"' '"+',"nombres") as paciente ,"turno" ,to_char("fechaIngreso",'+"'YYYY-MM-DD'"+'),to_char("fechaEvaluacion",'+"'YYYY-MM-DD'"+'),date_part('+"'year'"+',CURRENT_DATE)-date_part('+"'year'"+',fecha_nac),peso,talla,imc,"porcentajeCMB","porcentajeEPT","albSerica","ValGlobalSub" as "MIS","ingestaCalorica","ingestaProteica","diagNutricional","interveNutricional","obsNutricion","tipoPaciente"from public."Asistencial_nutricion" nut left join public."Asistencial_paciente" pac on nut.paciente_id=pac.id left join public."Asistencial_cas" cas on cas.id=pac.cas_id left join public. "Asistencial_usuario" usu on nut.usuario_id=usu.id where "fechaEvaluacion">= '+"'"+fecha+"'"

    cursor.execute(sql)
    resultados = cursor.fetchall()
    # Convertir la lista a JSON
    datos = []
    for fila in resultados:
        #print(fila)
        datos.append(dict(zip(('descripCas', 'nombre', 'fechaReg', 'paciente', 'turno', 'fechaIngreso', 'fechaEvaluacion', 'fecha_nac', 'peso', 'talla', 'imc', 'porcentajeCMB', 'porcentajeEPT', 'albSerica', 'MIS', 'ingestaCalorica', 'ingestaProteica', 'diagNutricional', 'interveNutricional', 'obsNutricion'), fila)))
    json_data = json.dumps(datos)

    return HttpResponse(json_data)

@api_view(['POST'])
@permission_required([IsAuthenticated])
def rep_Informe_Nutricion(request):
    body=request.body.decode('utf-8')
    objeto_python = json.loads(body)
    #Calculamos la fecha de hoy
    hoy = date.today()
    fecha_formateada = hoy.strftime("%m-%d-%Y")
    if objeto_python["fecha"]==None :
        fecha=fecha_formateada
    else :
        fecha=objeto_python["fecha"]
    # Importar las clases necesarias
    cursor = connection.cursor()
    sql = 'select "descripCas",usu."nombre",to_char("fechaReg",'+"'YYYY-MM-DD'"+') ,concat("ape_pat",'+"' '"+',"ape_mat",'+"' '"+',"nombres") as paciente ,"turno" ,to_char("fechaIngreso",'+"'YYYY-MM-DD'"+'),to_char("fechaEvaluacion",'+"'YYYY-MM-DD'"+'),date_part('+"'year'"+',CURRENT_DATE)-date_part('+"'year'"+',fecha_nac),peso,talla,imc,"porcentajeCMB","porcentajeEPT","albSerica","ValGlobalSub" as "MIS","ingestaCalorica","ingestaProteica","diagNutricional","interveNutricional","obsNutricion","tipoPaciente"from public."Asistencial_nutricion" nut left join public."Asistencial_paciente" pac on nut.paciente_id=pac.id left join public. "Asistencial_usuario" usu on nut.usuario_id=usu.id left join public."Asistencial_cas" cas on cas.id=usu.cas_id where "fechaEvaluacion">= '+"'"+fecha+"'"+' and "descripCas"= '+"'"+objeto_python["descripCas"]+"'"+' and usu.usuario='+"'"+objeto_python["usuario"]+"'"+'order by "fechaEvaluacion" desc'

    #where "fechaReg"='+"'"+fecha+"'"

    cursor.execute(sql)
    resultados = cursor.fetchall()
    # Convertir la lista a JSON
    datos = []
    for fila in resultados:
        #print(fila)
        datos.append(dict(zip(('descripCas', 'nombre', 'fechaReg', 'paciente', 'turno', 'fechaIngreso', 'fechaEvaluacion', 'fecha_nac', 'peso', 'talla', 'imc', 'porcentajeCMB', 'porcentajeEPT', 'albSerica', 'MIS', 'ingestaCalorica', 'ingestaProteica', 'diagNutricional', 'interveNutricional', 'obsNutricion'), fila)))
    json_data = json.dumps(datos)

    return HttpResponse(json_data)



@api_view(['POST'])
@permission_required([IsAuthenticated])
def rep_Asistencia_Pacientes(request):
    body=request.body.decode('utf-8')
    objeto_python = json.loads(body)
    #Calculamos la fecha de hoy
    hoy = date.today()
    fecha_formateada = hoy.strftime("%m-%d-%Y")

    if objeto_python["fecha"]==None :
        fecha=fecha_formateada
    else :
        fecha=objeto_python["fecha"]
    # Importar las clases necesarias
    cursor = connection.cursor()
    sql = 'select * from generar_reporte_asistencia(%s)'

    cursor.execute(sql, [fecha])

    # Obtener todos los resultados
    resultados = cursor.fetchall()
    # Convertir la lista a JSON
    datos = []
    for fila in resultados:
        #print(fila)
        datos.append(dict(zip(('descripCas', 'frecuencia','turno', 'tipoPuesto', 'sala','estadoAsistencia','observaFalta','fechaReg','paciente','numero','vigSeguro','num_doc','distrito'), fila)))
    json_data = json.dumps(datos)

    return HttpResponse(json_data)

@api_view(['POST'])
def webscrping(request):
    # Iniciar una sesión para mantener cookies
    session = requests.Session()

    # URL del login
    login_url = "http://sgss.essalud/sgss/servlet/hmainmenu"

    # Preparar el payload de login
    payload = {
        "W0005_USECOD": "44430347",
        "W0005_CLAVE": "Zhaetoo1234",
        "W0005BTNLOGIN": "Login",
    }

    # Realizar el POST al login
    response = session.post(login_url, data=payload)
    soup = BeautifulSoup(response.content, 'html.parser')
    print(soup)

    # Verificar si el login fue exitoso
    if "Centro Asistencial" in response.text.lower() or "paciente" in response.text.lower():
        print("✅ Login exitoso")
        
        # Ahora puedes navegar a otra URL interna
        datos_url = "http://sgss.essalud/sgss/servlet/pagina_interna"
        datos_response = session.get(datos_url)
        print(datos_response.text)

    else:
        print("❌ Login fallido o no autorizado")
        print(response.text)
    return {"error": response }

@api_view(['POST'])
@permission_required([IsAuthenticated])
def rep_cupos(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)
            
            # Obtener los valores de la solicitud y asegurarse de que sean del tipo esperado
            clinica_id = objeto_python.get("clinica_id")
            estado = objeto_python.get("estado")
            # Verificar si los valores son None o no están presentes
            

            # Importar las clases necesarias
            cursor = connection.cursor()
            
            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_reporte_cupos_detalles(%s, %s)'
            cursor.execute(sql, [clinica_id, estado])

            # Obtener todos los resultados
            resultados = cursor.fetchall()
            #print(resultados)
            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('dni', 'pacienteNombre','fecha','usuario','clinica','turno','frecuencia','parameCentroPuesto_id','paciente_id','distrito','numero_totales','numero_asignado','numero_liberado','tipoPuesto','serologia'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)

@api_view(['POST'])
@permission_required([IsAuthenticated])
def rep_cupos_agrupados(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)
            
            # Obtener los valores de la solicitud y asegurarse de que sean del tipo esperado
            clinica_id = objeto_python.get("clinica_id")
            tipo_puesto = objeto_python.get("tipo_puesto")
            # Verificar si los valores son None o no están presentes
            
            # Importar las clases necesarias
            cursor = connection.cursor()
            
            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_reporte_cupos(%s, %s)'
            cursor.execute(sql, [clinica_id, tipo_puesto])

            # Obtener todos los resultados
            resultados = cursor.fetchall()

            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('distrito','clinica', 'turno','frecuencia','tipoPuesto','numero_totales','numero_asignado','numero_liberado','cas_id','fecha'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)

@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_confirmacion_liberacion_cupo(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)
            
            # Obtener los valores de la solicitud y asegurarse de que sean del tipo esperado
            asigCuposPac_id = objeto_python.get("asigCuposPac_id")
            perfil = objeto_python.get("perfil")
            # Verificar si los valores son None o no están presentes
            

            # Importar las clases necesarias
            cursor = connection.cursor()
            
            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_confirmacion_liberacion_cupo(%s,%s)'
            cursor.execute(sql, [asigCuposPac_id,perfil])

            # Obtener todos los resultados
            resultados = cursor.fetchall()

            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('dias'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)

@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_reporte_inasistencia(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)
            
            # Obtener los valores de la solicitud y asegurarse de que sean del tipo esperado
            cas_id = objeto_python.get("cas_id")
            dias = objeto_python.get("dias")
            # Verificar si los valores son None o no están presentes
            

            # Importar las clases necesarias
            cursor = connection.cursor()
            
            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_reporte_inasistencia(%s,%s)'
            cursor.execute(sql, [cas_id,dias])

            # Obtener todos los resultados
            resultados = cursor.fetchall()

            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('paciente','frecuencia','turno','clinica','fecha','telefono','telefono_alterno','observacion'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)

@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_turno_actual(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)
            
            # Obtener los valores de la solicitud y asegurarse de que sean del tipo esperado
            paciente_id = objeto_python.get("paciente_id")

            # Importar las clases necesarias
            cursor = connection.cursor()
            
            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_turno_actual(%s)'
            cursor.execute(sql, [paciente_id])

            # Obtener todos los resultados
            resultados = cursor.fetchall()
            
            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('frecuencia','turno','numeroPuesto','tipoPuesto','asigCupoId','casId','descripCas'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)


@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_liberacion_cupo(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)
            
            # Obtener los valores de la solicitud y asegurarse de que sean del tipo esperado
            asigcupospac_id = objeto_python.get("asigcupospac_id")
            usuario = objeto_python.get("usuario")
            motivo = objeto_python.get("motivo")            

            # Importar las clases necesarias
            cursor = connection.cursor()
            
            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_liberacion_cupo(%s,%s,%s)'
            cursor.execute(sql, [asigcupospac_id,usuario,motivo])
            
            # Obtener todos los resultados
            resultados = cursor.fetchall()
            #print(resultados)
            # Convertir los resultados a formato JSON
            datos = []
            
            for fila in resultados:
                datos.append(dict(zip(('nombres'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)

def generar_asignacion_cupo(request):
    if request.method == 'POST':
        try:
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)

            # Obtener los valores del formulario
            fecha_reg = objeto_python.get("fecha_reg")
            usuario_reg= objeto_python.get("usuario_reg")
            paciente = objeto_python.get("paciente")
            parameCentroPuesto = objeto_python.get("parameCentroPuesto")
            estado = objeto_python.get("estado")
            # Importar las clases necesarias
            cursor = connection.cursor()
            
            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_asignacion_cupo(%s,%s,%s,%s)'
            cursor.execute(sql, [fecha_reg,usuario_reg,paciente,parameCentroPuesto])
            
            # Obtener todos los resultados
            resultados = cursor.fetchall()
            #print(resultados)
            # Convertir los resultados a formato JSON
            datos = []
            
            for fila in resultados:
                datos.append(dict(zip(('nombres'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)


def generar_resolucion_formulario(request):
    if request.method == 'POST':
        try:
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)

            # Obtener los valores del formulario
            id_formulario = objeto_python.get("id_formulario")
            respuesta= objeto_python.get("respuesta")
            usuario = objeto_python.get("usuario")
            fecha = objeto_python.get("fecha")

            # Importar las clases necesarias
            cursor = connection.cursor()
            
            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_resolucion_formulario(%s,%s,%s,%s)'
            cursor.execute(sql, [id_formulario,respuesta,usuario,fecha])
            
            # Obtener todos los resultados
            resultados = cursor.fetchall()
            #print(resultados)
            # Convertir los resultados a formato JSON
            datos = []
            
            for fila in resultados:
                datos.append(dict(zip(('nombres'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)


@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_liberacion_cupo_documento(request):
    if request.method == 'POST':
        try:
            # Obtener los valores del formulario
            asigcupospac_id = request.POST.get("asigcupospac_id")
            motivo_liberacion = request.POST.get("motivo_liberacion")
            usuario_reg_termino = request.POST.get("usuario_reg_termino")
            sustento = request.FILES.get("sustento")
            # Obtener la instancia del modelo y actualizar
            asigCuposPacItem = get_object_or_404(asigCuposPac, id=asigcupospac_id)
            asigCuposPacItem.sustento = sustento
            asigCuposPacItem.save()

            # Importar las clases necesarias
            cursor = connection.cursor()
            
            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_liberacion_cupo(%s,%s,%s)'
            cursor.execute(sql, [asigcupospac_id,usuario_reg_termino,motivo_liberacion])
            
            # Obtener todos los resultados
            resultados = cursor.fetchall()
            #print(resultados)
            # Convertir los resultados a formato JSON
            datos = []
            
            for fila in resultados:
                datos.append(dict(zip(('nombres'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)


def generar_buscar_cupo(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)
            
            # Obtener los valores de la solicitud y asegurarse de que sean del tipo esperado
            paciente_id = objeto_python.get("paciente_id")
            # Verificar si los valores son None o no están presentes
            

            # Importar las clases necesarias
            cursor = connection.cursor()
            
            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_buscar_cupo(%s)'
            cursor.execute(sql, [paciente_id])

            # Obtener todos los resultados
            resultados = cursor.fetchall()
            
            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('frecuencia'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)

@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_movimiento_paciente(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)
            
            # Obtener los valores de la solicitud y asegurarse de que sean del tipo esperado
            id_asignacioncupo = objeto_python.get("id_asignacioncupo")
            tipo_movimiento=objeto_python.get("tipo_movimiento")
            id_parameCentroPuesto = objeto_python.get("id_parameCentroPuesto")
            usuario = objeto_python.get("usuario")
            fecha = objeto_python.get("fecha")
            cas = objeto_python.get("cas")
            # Verificar si los valores son None o no están presentes
            

            # Importar las clases necesarias
            cursor = connection.cursor()
            
            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_movimiento_paciente(%s,%s,%s,%s,%s)'
            cursor.execute(sql, [id_asignacioncupo,tipo_movimiento,id_parameCentroPuesto,usuario,fecha])

            # Obtener todos los resultados
            resultados = cursor.fetchall()

            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('frecuencia'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"er0ror": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)

@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_lista_disponible(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)
            
            # Obtener los valores de la solicitud y asegurarse de que sean del tipo esperado
            clinica_id = objeto_python.get("clinica_id")
            # Verificar si los valores son None o no están presentes

            # Importar las clases necesarias
            cursor = connection.cursor()
            
            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_lista_disponible(%s)'
            cursor.execute(sql, [clinica_id])

            # Obtener todos los resultados
            resultados = cursor.fetchall()

            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('turno','frecuencia','tipoPuesto','numeroPuesto','id_ParametroCentro'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)

@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_lista_hospital(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)
            
            # Obtener los valores de la solicitud y asegurarse de que sean del tipo esperado
            # Verificar si los valores son None o no están presentes

            # Importar las clases necesarias
            cursor = connection.cursor()
            
            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_lista_hospital()'
            cursor.execute(sql, [])

            # Obtener todos los resultados
            resultados = cursor.fetchall()

            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('idHospital','hospital','ruc','tipo_institucion','direccion','inicio_actividades','estado'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)

@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_lista_medico(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)
            
            # Obtener los valores de la solicitud y asegurarse de que sean del tipo esperado
            # Verificar si los valores son None o no están presentes

            # Importar las clases necesarias
            cursor = connection.cursor()
            
            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_lista_medico()'
            cursor.execute(sql, [])

            # Obtener todos los resultados
            resultados = cursor.fetchall()

            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('idMedico','nombre','dni','correo','telefono'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)

@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_lista_pacientes(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)
            
            # Obtener los valores de la solicitud y asegurarse de que sean del tipo esperado
            clinica_id = objeto_python.get("clinica_id")
            # Verificar si los valores son None o no están presentes

            # Importar las clases necesarias
            cursor = connection.cursor()
            
            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_lista_pacientes(%s)'
            cursor.execute(sql, [clinica_id])

            # Obtener todos los resultados
            resultados = cursor.fetchall()

            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('nombreCompleto','paciente_id','dni'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)

def carga_masiva_paciente(request):
    if request.method == 'POST' and request.FILES['file']:
        file = request.FILES['file']
        try:
            # Por ejemplo, puedes iterar sobre las filas y guardar cada una como un registro en la base de datos
            workbook = openpyxl.load_workbook(file, read_only=True, data_only=True)
            sheet = workbook.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                tipo_doc, num_doc, ape_pat, ape_mat, nombres, fecha_nac, sexo, hospital_id,medico_id,coordenadas,direccion,telefono,ubigeo_id,serologia = row
                # Verificar si el paciente ya existe en la base de datos
                paciente_existente = paciente.objects.filter(num_doc=num_doc).first()
                srid = 4326
                if paciente_existente:
                    # Actualizar los datos del paciente existente
                    paciente_existente.tipo_doc = maestro.objects.get(id=tipo_doc)
                    paciente_existente.num_doc = num_doc
                    paciente_existente.ape_pat = ape_pat
                    paciente_existente.ape_mat = ape_mat
                    paciente_existente.nombres = nombres
                    paciente_existente.fecha_nac = fecha_nac
                    paciente_existente.sexo = sexo
                    paciente_existente.hospital_id = hospital.objects.get(id=hospital_id)
                    paciente_existente.medico_id = medico.objects.get(id=medico_id)
                    paciente_existente.latitud = coordenadas.split(',')[0].strip()
                    paciente_existente.longitud = coordenadas.split(',')[1].strip()
                    paciente_existente.direccion = direccion
                    paciente_existente.telefono = telefono
                    paciente_existente.ubigeo_id = ubigeo.objects.get(id=ubigeo_id)
                    paciente_existente.serologia = serologia
                    paciente_existente.save()
                else:
                    # Crear un nuevo registro de paciente
                    paciente_nuevo = paciente(
                        tipo_doc=maestro.objects.get(id=tipo_doc),
                        num_doc=num_doc,
                        ape_pat=ape_pat,
                        ape_mat=ape_mat,
                        nombres=nombres,
                        fecha_nac=fecha_nac,
                        sexo=sexo,
                        hospital_id=hospital.objects.get(id=hospital_id),
                        medico_id=medico.objects.get(id=medico_id),
                        latitud=coordenadas.split(',')[0].strip(),
                        longitud=coordenadas.split(',')[1].strip(),
                        direccion=direccion,
                        telefono=telefono,
                        ubigeo_id=ubigeo.objects.get(id=ubigeo_id),
                        serologia = serologia
                    )
                    paciente_nuevo.save()

            return JsonResponse({'message': 'Datos guardados exitosamente.'})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    else:
        return JsonResponse({'error': 'Debes enviar un archivo.'}, status=400)

def parse_date(date_str):
    """Convert date from DD/MM/YYYY to YYYY-MM-DD"""
    try:
        return datetime.strptime(date_str, "%d/%m/%Y").date()
    except ValueError:
        return None  # or handle the error as needed

def carga_masiva_laboratorio(request):
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        
        # Verificar la extensión del archivo
        if not file.name.endswith('.txt'):
            return HttpResponse('El archivo no tiene la extensión .txt')

        try:
            file_data = file.read().decode('utf-8')
            lines = file_data.splitlines()
            # Asumiendo que la primera línea es el encabezado
            for line in lines[1:]:
                fields = line.split('|')
                centro, periodo, area, codserv, servicio, codact, actividad, codsubact, subactividad, acto_medico, fecha_atencion, fecha_solicitud, fecha_cita, fecha_resultado, num_solicitud, dni_solicita, prof_solicita, tipexamen, arealab, sede, examen, descexamen, dni_profesional, profesional, dni_paciente, h_c, paciente, telefonos, annos, meses, dias, sexo, redasist, tipo_paciente, cas_adscripcion, desc_adscripcion, diagnostico, des_diagn, tip_diagn, resultado, tipo_resultado, categoria_resul, fecha_registro, usuario_registro, usario_modifica, fecha_modifica, centro_origen_solicitud, descentorisol, hora_registro, autogenerado, ubigeo_domic, desc_ubigeo, domicilio, desc_topico, turno, tipestudio, tipo_gravidez, muestra, orden_plantilla, desc_plantilla, unidadvalor, valor_resultado, informe_resultado, observresultado, otrosvalores = fields
                valor_texto=valor_resultado
                valor_resultado=re.sub(r'[^\d.]+', '', valor_resultado)

                if valor_resultado == "":
                    valor_resultado = None
                # Verificar si el paciente ya existe en la base de datos
                laboratorio_existente = laboratorio.objects.filter(subactividad=subactividad,fecha_solicitud=parse_date(fecha_solicitud),fecha_resultado=parse_date(fecha_resultado),num_solicitud=num_solicitud,desc_plantilla=desc_plantilla,dni_paciente=dni_paciente,unidad=unidadvalor,valor=valor_resultado,valor_otros=otrosvalores,informe_resultado=informe_resultado).first()

                if laboratorio_existente:
                    if valor_resultado == "":
                        valor_resultado = None
                    # Actualizar los datos del paciente existente
                    laboratorio_existente.subactividad = subactividad
                    laboratorio_existente.fecha_toma_muestra = parse_date(fecha_atencion)
                    laboratorio_existente.fecha_solicitud = parse_date(fecha_solicitud)
                    laboratorio_existente.fecha_resultado = parse_date(fecha_resultado)
                    laboratorio_existente.dni_paciente = dni_paciente
                    laboratorio_existente.num_solicitud = num_solicitud
                    laboratorio_existente.desc_plantilla = desc_plantilla
                    laboratorio_existente.unidad = unidadvalor
                    laboratorio_existente.valor = valor_resultado
                    laboratorio_existente.valor_otros = otrosvalores
                    laboratorio_existente.desc_examen = descexamen
                    laboratorio_existente.informe_resultado = informe_resultado
                    laboratorio_existente.autogenerado=autogenerado
                    laboratorio_existente.valor_texto=valor_texto
                    laboratorio_existente.save()
                else:
                    if valor_resultado == "":
                        valor_resultado = None
                    # Crear un nuevo registro de paciente
                    laboratorio_nuevo = laboratorio(
                        subactividad=subactividad,
                        fecha_toma_muestra=parse_date(fecha_atencion),
                        fecha_solicitud=parse_date(fecha_solicitud),
                        fecha_resultado=parse_date(fecha_resultado),
                        num_solicitud=num_solicitud,
                        dni_paciente=dni_paciente,
                        desc_plantilla=desc_plantilla,
                        unidad=unidadvalor,
                        valor=valor_resultado,
                        valor_otros=otrosvalores,
                        desc_examen = descexamen,
                        informe_resultado = informe_resultado,
                        autogenerado=autogenerado,
                        valor_texto=valor_texto
                    )
                    laboratorio_nuevo.save()
                    

            return JsonResponse({'message': 'Datos guardados exitosamente.'})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    else:
        return JsonResponse({'error': 'Debes enviar un archivo.'}, status=400)
    """ if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        if not file.name.endswith('.txt'):
            return HttpResponse('El archivo no tiene la extensión .txt')

        # Definir la ruta temporal en la cual guardar el archivo
        temp_dir = 'C:/Program Files/PostgreSQL/10/data'
        file_path = os.path.join(temp_dir, file.name)

        try:
            # Crear la carpeta temporal si no existe
            os.makedirs(temp_dir, exist_ok=True)

            # Guardar temporalmente el archivo en el servidor
            with open(file_path, 'wb+') as destination:
                for chunk in file.chunks():
                    destination.write(chunk)

            # Llamar al procedimiento almacenado en PostgreSQL usando CALL
            with connection.cursor() as cursor:
                cursor.execute(f"CALL generar_importar_datos_laboratorio(%s)", [file_path])

            # Eliminar el archivo temporal después de la ejecución
            os.remove(file_path)

            return JsonResponse({'message': 'Datos guardados exitosamente.'})

        except Exception as e:
            # Manejo de errores, eliminar el archivo si hubo un problema
            if os.path.exists(file_path):
                os.remove(file_path)
            return JsonResponse({'error': str(e)}, status=500)
    else:
        return JsonResponse({'error': 'Debes enviar un archivo.'}, status=400)
 """

def carga_masiva_laboratorio_temp(request):
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        
        # Verificar la extensión del archivo
        if not file.name.endswith('.txt'):
            return HttpResponse('El archivo no tiene la extensión .txt')

        try:
            file_data = file.read().decode('utf-8')
            lines = file_data.splitlines()
            # Asumiendo que la primera línea es el encabezado
            for line in lines[1:]:
                fields = line.split('|')
                centro, periodo, area, codserv, servicio, codact, actividad, codsubact, subactividad, acto_medico, fecha_atencion, fecha_solicitud, fecha_cita, fecha_resultado, num_solicitud, dni_solicita, prof_solicita, tipexamen, arealab, sede, examen, descexamen, dni_profesional, profesional, dni_paciente, h_c, paciente, telefonos, annos, meses, dias, sexo, redasist, tipo_paciente, cas_adscripcion, desc_adscripcion, diagnostico, des_diagn, tip_diagn, resultado, tipo_resultado, categoria_resul, fecha_registro, usuario_registro, usario_modifica, fecha_modifica, centro_origen_solicitud, descentorisol, hora_registro, autogenerado, ubigeo_domic, desc_ubigeo, domicilio, desc_topico, turno, tipestudio, tipo_gravidez, muestra, orden_plantilla, desc_plantilla, unidadvalor, valor_resultado, informe_resultado, observresultado, otrosvalores = fields
                
                valor_texto=valor_resultado
                valor_resultado=re.sub(r'[^\d.]+', '', valor_resultado)

                """ if valor_resultado == "":
                    valor_resultado = None """

                """ if laboratorio_existente:
                    if valor_resultado == "":
                        valor_resultado = None
                    # Actualizar los datos del paciente existente
                    laboratorio_existente.subactividad = subactividad
                    laboratorio_existente.fecha_toma_muestra = parse_date(fecha_atencion)
                    laboratorio_existente.fecha_solicitud = parse_date(fecha_solicitud)
                    laboratorio_existente.fecha_resultado = parse_date(fecha_resultado)
                    laboratorio_existente.dni_paciente = dni_paciente
                    laboratorio_existente.num_solicitud = num_solicitud
                    laboratorio_existente.desc_plantilla = desc_plantilla
                    laboratorio_existente.unidad = unidadvalor
                    laboratorio_existente.valor = valor_resultado
                    laboratorio_existente.valor_otros = otrosvalores
                    laboratorio_existente.desc_examen = descexamen
                    laboratorio_existente.informe_resultado = informe_resultado
                    laboratorio_existente.autogenerado=autogenerado
                    laboratorio_existente.valor_texto=valor_texto
                    laboratorio_existente.save()
                else: """
                if valor_resultado == "":
                    valor_resultado = None
                laboratorio_nuevo = laboratorioTemp(
                    subactividad=subactividad,
                    fecha_atencion=parse_date(fecha_atencion),
                    fecha_solicitud=parse_date(fecha_solicitud),
                    fecha_resultado=parse_date(fecha_resultado),
                    num_solicitud=num_solicitud,
                    dni_paciente=dni_paciente,
                    desc_plantilla=desc_plantilla,
                    unidadvalor=unidadvalor,
                    valor_resultado=valor_resultado,
                    otrosvalores=otrosvalores,
                    descexamen = descexamen,
                    informe_resultado = informe_resultado,
                    autogenerado=autogenerado
                )
                laboratorio_nuevo.save()
                    

            return JsonResponse({'message': 'Datos guardados exitosamente.'})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    else:
        return JsonResponse({'error': 'Debes enviar un archivo.'}, status=400)

def carga_masiva_lista_espera(request):
        if request.method == 'POST' and request.FILES.get('file'):
            file = request.FILES['file']
        
            # Verificar la extensión del archivo
            if not file.name.endswith('.txt'):
                return HttpResponse('El archivo no tiene la extensión .txt')

            try:
                file_data = file.read().decode('utf-8')
                lines = file_data.splitlines()
                    
                nuevos_laboratorios = []  # Lista para acumular nuevos objetos
                actualizaciones = []  # Lista para objetos a actualizar

                # Asumiendo que la primera línea es el encabezado
                for line in lines[1:]:
                        fields = line.split('|')
                        # Asignar valores a las variables correspondientes...
                        # Código similar al tuyo para parsear las líneas y asignar los valores...
                        
                        # Procesar el valor del resultado
                        valor_texto = valor_resultado
                        valor_resultado = re.sub(r'[^\d.]+', '', valor_resultado)
                        if valor_resultado == "":
                            valor_resultado = None
                        
                        # Verificar si el paciente ya existe en la base de datos
                        laboratorio_existente = laboratorio.objects.filter(
                            subactividad=subactividad,
                            fecha_solicitud=parse_date(fecha_solicitud),
                            fecha_resultado=parse_date(fecha_resultado),
                            num_solicitud=num_solicitud,
                            desc_plantilla=desc_plantilla,
                            dni_paciente=dni_paciente,
                            unidad=unidadvalor,
                            valor=valor_resultado,
                            valor_otros=otrosvalores,
                            informe_resultado=informe_resultado
                        ).first()

                        if laboratorio_existente:
                            # Actualizar los datos del paciente existente
                            laboratorio_existente.subactividad = subactividad
                            laboratorio_existente.fecha_toma_muestra = parse_date(fecha_atencion)
                            laboratorio_existente.fecha_solicitud = parse_date(fecha_solicitud)
                            laboratorio_existente.fecha_resultado = parse_date(fecha_resultado)
                            laboratorio_existente.dni_paciente = dni_paciente
                            laboratorio_existente.num_solicitud = num_solicitud
                            laboratorio_existente.desc_plantilla = desc_plantilla
                            laboratorio_existente.unidad = unidadvalor
                            laboratorio_existente.valor = valor_resultado
                            laboratorio_existente.valor_otros = otrosvalores
                            laboratorio_existente.desc_examen = descexamen
                            laboratorio_existente.informe_resultado = informe_resultado
                            laboratorio_existente.autogenerado = autogenerado
                            laboratorio_existente.valor_texto = valor_texto
                            actualizaciones.append(laboratorio_existente)  # Añadir a la lista de actualizaciones
                        else:
                            # Crear un nuevo registro de paciente
                            laboratorio_nuevo = laboratorio(
                                subactividad=subactividad,
                                fecha_toma_muestra=parse_date(fecha_atencion),
                                fecha_solicitud=parse_date(fecha_solicitud),
                                fecha_resultado=parse_date(fecha_resultado),
                                num_solicitud=num_solicitud,
                                dni_paciente=dni_paciente,
                                desc_plantilla=desc_plantilla,
                                unidad=unidadvalor,
                                valor=valor_resultado,
                                valor_otros=otrosvalores,
                                desc_examen=descexamen,
                                informe_resultado=informe_resultado,
                                autogenerado=autogenerado,
                                valor_texto=valor_texto
                            )
                            nuevos_laboratorios.append(laboratorio_nuevo)  # Añadir a la lista de nuevos objetos

                    # Utilizar bulk_create para nuevos objetos
                if nuevos_laboratorios:
                        laboratorio.objects.bulk_create(nuevos_laboratorios, batch_size=1000)
                    
                    # Guardar actualizaciones de manera más eficiente
                if actualizaciones:
                        with transaction.atomic():  # Usar una transacción para guardar en bloque
                            for lab in actualizaciones:
                                lab.save()

                return JsonResponse({'message': 'Datos guardados exitosamente.'})
            except Exception as e:
                return JsonResponse({'error': str(e)}, status=500)
        else:
            return JsonResponse({'error': 'Debes enviar un archivo.'}, status=400)

@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_lista_pacientes_nuevos(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)
            
            # Obtener los valores de la solicitud y asegurarse de que sean del tipo esperado
            

            # Verificar si los valores son None o no están presentes

            # Importar las clases necesarias
            cursor = connection.cursor()
            
            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_lista_pacientes_nuevos()'
            cursor.execute(sql)

            # Obtener todos los resultados
            resultados = cursor.fetchall()

            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('paciente','id','num_doc'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)

@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_lista_cupos(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)

            # Verificar si los valores son None o no están presentes

            # Importar las clases necesarias
            cursor = connection.cursor()
            
            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_lista_cupos()'
            cursor.execute(sql)

            # Obtener todos los resultados
            resultados = cursor.fetchall()

            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('frecuencia','turno','tipoPuesto','numeroPuesto','descripCas','estado','idCupo','distrito'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)

@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_actualizar_cupo(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)

            # Verificar si los valores son None o no están presentes
            id_parameCentroPuesto = objeto_python.get("id_parameCentroPuesto")
            estado = objeto_python.get("estado")

            # Importar las clases necesarias
            cursor = connection.cursor()
            
            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_actualizar_cupo(%s,%s)'
            cursor.execute(sql, [id_parameCentroPuesto,estado])

            # Obtener todos los resultados
            resultados = cursor.fetchall()

            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('frecuencia','turno','tipoPuesto'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)


@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_lista_zona(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')


            # Verificar si los valores son None o no están presentes
            
            # Importar las clases necesarias
            cursor = connection.cursor()
            
            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_lista_zona()'
            cursor.execute(sql, [])

            # Obtener todos los resultados
            resultados = cursor.fetchall()

            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('id_zona','zona'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)


@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_lista_ipress(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)

            # Verificar si los valores son None o no están presentes
            tipoCas = objeto_python.get("tipoCas")
            # Importar las clases necesarias
            cursor = connection.cursor()
            
            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_lista_ipress(%s)'
            cursor.execute(sql, [tipoCas])

            # Obtener todos los resultados
            resultados = cursor.fetchall()

            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('id_cas','descripCas','distrito','cordeCas','latitud','longitud'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)

@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_lista_distrito(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)

            # Verificar si los valores son None o no están presentes
            # Importar las clases necesarias
            cursor = connection.cursor()
            
            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_lista_distrito()'
            cursor.execute(sql, [])

            # Obtener todos los resultados
            resultados = cursor.fetchall()

            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('ubigeo_reniec','departamento','provincia','distrito','latitud','longitud'), fila)))
            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)

@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_lista_geolocalizacion(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)

            # Verificar si los valores son None o no están presentes
            # Importar las clases necesarias
            cursor = connection.cursor()

            ubigeo = objeto_python.get("ubigeo")
            clinica = objeto_python.get("clinica")
            documento = objeto_python.get("documento")

            if not clinica:
                clinica = None
            if not documento:
                documento = None
            if not ubigeo:
                ubigeo = None
            # Definir la consulta SQL y ejecutarla
            #sql = 'select * from generar_lista_geolocalizacion(%s,%s,%s)'%s::bigint, %s::character varying[], %s::character varying[]
            sql = 'select * from generar_lista_geolocalizacion(%s::bigint[],%s::bigint[],%s::character varying[])'
            cursor.execute(sql, [ubigeo,clinica,documento])

            # Obtener todos los resultados
            resultados = cursor.fetchall()

            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('idpaciente','num_doc','ape_pat','ape_mat','nombres','paciente','direccion','referencia','telefono','telefono_alterno','cordepac','latitud','longitud','distritopaciente','turno','frecuencia','idclinica','descripcas','distritoclinica'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)

@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_lista_geolocalizacion_jaen(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)

            # Verificar si los valores son None o no están presentes
            # Importar las clases necesarias
            cursor = connection.cursor()

            ubigeo = objeto_python.get("ubigeo")
            clinica = objeto_python.get("clinica")
            documento = objeto_python.get("documento")

            if not clinica:
                clinica = None
            if not documento:
                documento = None
            if not ubigeo:
                ubigeo = None
            # Definir la consulta SQL y ejecutarla
            #sql = 'select * from generar_lista_geolocalizacion(%s,%s,%s)'%s::bigint, %s::character varying[], %s::character varying[]
            sql = 'select * from generar_lista_geolocalizacion_jaen(%s::bigint[],%s::bigint[],%s::character varying[])'
            cursor.execute(sql, [ubigeo,clinica,documento])

            # Obtener todos los resultados
            resultados = cursor.fetchall()

            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('idpaciente','num_doc','ape_pat','ape_mat','nombres','paciente','direccion','referencia','telefono','telefono_alterno','cordepac','latitud','longitud','distritopaciente','turno','frecuencia','idclinica','descripcas','distritoclinica'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)

@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_lista_paciente(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)

            # Verificar si los valores son None o no están presentes
            # Importar las clases necesarias
            cursor = connection.cursor()
            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_lista_pacientes()'
            cursor.execute(sql, [])

            # Obtener todos los resultados
            resultados = cursor.fetchall()

            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('paciente','num_doc'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)

@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_calculo_numero_cuidadores(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)

            # Verificar si los valores son None o no están presentes
            paciente = objeto_python.get("paciente")
            # Importar las clases necesarias
            cursor = connection.cursor()
            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_calculo_numero_cuidadores(%s)'
            cursor.execute(sql, [paciente])

            # Obtener todos los resultados
            resultados = cursor.fetchall()

            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('numero'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)

@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_buscar_chatbot(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)

            # Verificar si los valores son None o no están presentes
            dni = objeto_python.get("dni")
            # Importar las clases necesarias
            cursor = connection.cursor()
            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_buscar_chatbot(%s)'
            cursor.execute(sql, [dni])

            # Obtener todos los resultados
            resultados = cursor.fetchall()

            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('estado'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)

@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_busqueda_pacientes(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)

            # Verificar si los valores son None o no están presentes
            dni = objeto_python.get("dni")
            
            # Importar las clases necesarias
            cursor = connection.cursor()

            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_busqueda_pacientes(%s)'
            cursor.execute(sql, [dni])

            # Obtener todos los resultados
            resultados = cursor.fetchall()

            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('id_paciente','nombre_paciente','ape_pat','ape_mat','latitud','longitud','direccion','telefono','telefonoAlterno','ubigeo_id','departamento','provincia','distrito','clinica','turno','frecuencia','tipoPuesto'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)


@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_lista_geolocalizacion_mapa(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)
            
            # Importar las clases necesarias
            cursor = connection.cursor()

            ubigeo = objeto_python.get("ubigeo")
            clinica = objeto_python.get("clinica")
            documento = objeto_python.get("documento")

            if not clinica:
                clinica = None
            if not documento:
                documento = None
            if not ubigeo:
                ubigeo = None

            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_lista_geolocalizacion_mapa(%s::bigint[],%s::bigint[],%s::character varying[])'
            cursor.execute(sql, [ubigeo,clinica,documento])

            # Obtener todos los resultados
            resultados = cursor.fetchall()

            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('idClinica','latitud','longitud','nombreClinica','num_asignado','num_liberado','num_totales'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)

@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_lista_geolocalizacion_mapa_jaen(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)
            
            # Importar las clases necesarias
            cursor = connection.cursor()

            ubigeo = objeto_python.get("ubigeo")
            clinica = objeto_python.get("clinica")
            documento = objeto_python.get("documento")

            if not clinica:
                clinica = None
            if not documento:
                documento = None
            if not ubigeo:
                ubigeo = None

            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_lista_geolocalizacion_mapa_jaen(%s::bigint[],%s::bigint[],%s::character varying[])'
            cursor.execute(sql, [ubigeo,clinica,documento])

            # Obtener todos los resultados
            resultados = cursor.fetchall()

            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('idClinica','latitud','longitud','nombreClinica','num_asignado','num_liberado','num_totales'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)

@api_view(['POST'])
@permission_required([IsAuthenticated])
def reporte_asistencia_grafico(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)
            
            # Obtener los valores de la solicitud y asegurarse de que sean del tipo esperado
            periodo = objeto_python.get("periodo")
            mes = objeto_python.get("mes")
            # Verificar si los valores son None o no están presentes
            

            # Importar las clases necesarias
            cursor = connection.cursor()
            
            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_reporte_asistencia_grafico(%s, %s)'
            cursor.execute(sql, [periodo, mes])

            # Obtener todos los resultados
            resultados = cursor.fetchall()

            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('paciente'
				  ,'dia_1' 
				  ,'dia_2' 
				  ,'dia_3' 
				  ,'dia_4' 
				  ,'dia_5' 
				  ,'dia_6' 
				  ,'dia_7' 
				  ,'dia_8' 
				  ,'dia_9' 
				  ,'dia_10' 
				  ,'dia_11' 
				  ,'dia_12' 
				  ,'dia_13' 
				  ,'dia_14' 
				  ,'dia_15' 
				  ,'dia_16' 
				  ,'dia_17' 
				  ,'dia_18' 
				  ,'dia_19' 
				  ,'dia_20' 
				  ,'dia_21' 
				  ,'dia_22' 
				  ,'dia_23' 
				  ,'dia_24' 
				  ,'dia_25' 
				  ,'dia_26' 
				  ,'dia_27' 
				  ,'dia_28' 
				  ,'dia_29' 
				  ,'dia_30' 
				  ,'dia_31' 
				  ,'total_falto' 
				  ,'total_asistio'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)

@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_reporte_laboratorio(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)

            # Obtener los valores de la solicitud y asegurarse de que sean del tipo esperado
            periodo = objeto_python.get("periodo")
            mes = objeto_python.get("mes")
            
            # Importar las clases necesarias
            cursor = connection.cursor()
            
            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_reporte_laboratorio(%s,%s)'
            cursor.execute(sql,[mes,periodo])

            # Obtener todos los resultados
            resultados = cursor.fetchall()

            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('FECHA_RESULTADO', 'NUM_DOC', 'AUTOGENERADO', 'PACIENTE', 'SALA','ALBUMINA', 'CALCIO_TOTAL', 'CALCIO_POST', 'CALCIO_PRE', 'CALCIO_IONIZADO', 'CLORURO', 'COLESTEROL', 'CREATININA', 
                'CREATININA_POST', 'CREATININA_PRE', 'FERRITINA', 'GLUCOSA', 'FIERRO', 'COLESTEROL_HDL', 'COLESTEROL_LDL', 'PARATOHORMONA', 'FOSFATASA_ALCALINA', 'FOSFORO','FOSFORO_POST', 'FOSFORO_PRE', 'POTASIO', 'SODIO', 
                'TRANSAMINASA_GLUTAMICO_PIRUVIC', 'TRANSAMINASA_GLUTAMICO_OXALACE', 'TRANSFERRINA', 'TRIGLICERIDOS', 'UREA','ACIDO_URICO', 'GLOBULOS_BLANCOS', 'GLOBULOS_ROJOS', 'HEMATOCRITO', 'HEMOGLOBINA', 'PROTEINA_C_REACTIVA',
                'ANTIGENO_HEPATIT', 'ANTICUERPO_HEPATITIS_CORE', 'ANTICUERPO_HEPATITIS_C','UREA_PRE','UREA_POST','HCO3','ANTICUERPO_HEPATITIS_B','ANTIGENO_SUPERFICIE_HEPAT','ANTICUERPO_EPSILON','ACIDO_FOLICO','VITAMINA_B12','PROTEINAS',
                'SEROLOGIA','COD_IPRESS','IPRESS','TGO','ABASTONADOS','BASOFILOS','CONC_HB_CORPUSCULAR_MEDIA','EOSINOFILOS','HB_CORPUSCULAR_MEDIA','LINFOCITOS','MONOCITOS','NRCB','PCT','PDW','RDW_CV','RDW_SD','RECUENTO_PLAQUETAS',
                'SEGMENTADOS','VOLUMEN_CORPUSCULAR_MEDIO','VOLUMEN_PLAQUETARIO_MEDIO','PARATHORMONA_INTACTA','KTV', 'ANTICUERPO_HIV','POR_SATURACION','BETA_2','VITAMINA_D','TRU'
                ), fila)))
            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)

@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_reporte_protocolo(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)

            # Obtener los valores de la solicitud y asegurarse de que sean del tipo esperado
            
            fecha = objeto_python.get("fecha")
            
            # Importar las clases necesarias
            cursor = connection.cursor()
            
            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_reporte_laboratorio_protocolo(%s)'
            cursor.execute(sql,[fecha])

            # Obtener todos los resultados
            resultados = cursor.fetchall()

            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('FECHA_RESULTADO','NUM_DOC','paciente','HB_1','HB_2','HB_3','HB_4','HB_5','HB_6','FERRITINA','TRASFERRINA','HIERRO_1','HIERRO_2','HIERRO_3','HIERRO_4','HIERRO_5','HIERRO_6'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)


@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_reporte_laboratorio_protocolo_hemoglobina(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)

            # Obtener los valores de la solicitud y asegurarse de que sean del tipo esperado
            
            autogenerado = objeto_python.get("autogenerado")
            fecha = objeto_python.get("fecha")
            # Importar las clases necesarias
            cursor = connection.cursor()
            
            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_reporte_laboratorio_protocolo_hemoglobina(%s,%s)'
            cursor.execute(sql,[autogenerado,fecha])

            # Obtener todos los resultados
            resultados = cursor.fetchall()

            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('FECHA_RESULTADO','VALOR'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)

@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_reporte_laboratorio_protocolo_tgp(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)

            # Obtener los valores de la solicitud y asegurarse de que sean del tipo esperado
            
            autogenerado = objeto_python.get("autogenerado")
            fecha = objeto_python.get("fecha")
            # Importar las clases necesarias
            cursor = connection.cursor()
            
            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_reporte_laboratorio_protocolo_tgp(%s,%s)'
            cursor.execute(sql,[autogenerado,fecha])

            # Obtener todos los resultados
            resultados = cursor.fetchall()

            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('FECHA_RESULTADO','VALOR'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)


@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_reporte_laboratorio_protocolo_dosis_epo(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)

            # Obtener los valores de la solicitud y asegurarse de que sean del tipo esperado
            
            autogenerado = objeto_python.get("autogenerado")
            fecha = objeto_python.get("fecha")
            # Importar las clases necesarias
            cursor = connection.cursor()
            
            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_reporte_laboratorio_protocolo_dosis_epo(%s,%s)'
            cursor.execute(sql,[autogenerado,fecha])

            # Obtener todos los resultados
            resultados = cursor.fetchall()

            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('FECHA_RESULTADO','VALOR'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)

@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_reporte_laboratorio_protocolo_calcio(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)

            # Obtener los valores de la solicitud y asegurarse de que sean del tipo esperado
            
            autogenerado = objeto_python.get("autogenerado")
            fecha = objeto_python.get("fecha")
            # Importar las clases necesarias
            cursor = connection.cursor()
            
            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_reporte_laboratorio_protocolo_calcio(%s,%s)'
            cursor.execute(sql,[autogenerado,fecha])

            # Obtener todos los resultados
            resultados = cursor.fetchall()

            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('FECHA_RESULTADO','VALOR'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)


@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_reporte_laboratorio_protocolo_paratohormona(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)

            # Obtener los valores de la solicitud y asegurarse de que sean del tipo esperado
            
            autogenerado = objeto_python.get("autogenerado")
            fecha = objeto_python.get("fecha")
            # Importar las clases necesarias
            cursor = connection.cursor()
            
            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_reporte_laboratorio_protocolo_paratohormona(%s,%s)'
            cursor.execute(sql,[autogenerado,fecha])

            # Obtener todos los resultados
            resultados = cursor.fetchall()

            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('FECHA_RESULTADO','VALOR'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)


@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_reporte_laboratorio_protocolo_fosforo(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)

            # Obtener los valores de la solicitud y asegurarse de que sean del tipo esperado
            
            autogenerado = objeto_python.get("autogenerado")
            fecha = objeto_python.get("fecha")
            # Importar las clases necesarias
            cursor = connection.cursor()
            
            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_reporte_laboratorio_protocolo_fosforo(%s,%s)'
            cursor.execute(sql,[autogenerado,fecha])

            # Obtener todos los resultados
            resultados = cursor.fetchall()

            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('FECHA_RESULTADO','VALOR'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)

@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_reporte_laboratorio_protocolo_ktv(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)

            # Obtener los valores de la solicitud y asegurarse de que sean del tipo esperado
            
            dni = objeto_python.get("dni")
            fecha = objeto_python.get("fecha")
            # Importar las clases necesarias
            cursor = connection.cursor()
            
            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_reporte_laboratorio_protocolo_ktv(%s,%s)'
            cursor.execute(sql,[dni,fecha])

            # Obtener todos los resultados
            resultados = cursor.fetchall()

            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('FECHA_RESULTADO','VALOR'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)

@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_reporte_cupos_por_tipo_detalles(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)

            # Verificar si los valores son None o no están presentes
            clinica_id = objeto_python.get("clinica_id")
            motivo = objeto_python.get("motivo")
            # Importar las clases necesarias
            cursor = connection.cursor()
            
            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_reporte_cupos_por_tipo_detalles(%s,%s)'
            cursor.execute(sql, [clinica_id,motivo])

            # Obtener todos los resultados
            resultados = cursor.fetchall()

            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('dni', 'pacienteNombre','fecha','usuario','clinica','turno','frecuencia','parameCentroPuesto_id','paciente_id','distrito','numero_totales','numero_asignado','numero_liberado','tipoPuesto','edad','frecuencia_nuevo','turno_nuevo','tipo_puesto_nuevo','hospital','fecha_efectiva'), fila)))
            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)

@api_view(['POST'])
@permission_required([IsAuthenticated])
def actualizar_tipo_paciente(request):
    if request.method == 'POST':
        try:
            # Importar las clases necesarias y definir la consulta SQL
            with connection.cursor() as cursor:
                sql = 'CALL public.actualizar_tipo_paciente()'
                cursor.execute(sql)

            # Devolver la respuesta JSON
            return JsonResponse("Exito", safe=False)

        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)


@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_reporte_nuevos_reingresos(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)

            # Importar las clases necesarias
            cursor = connection.cursor()
            
            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_reporte_nuevos_reingresos()'
            cursor.execute(sql, [])

            # Obtener todos los resultados
            resultados = cursor.fetchall()

            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('fecha_reg', 'num_doc','paciente','tipo_seguro','direccion','distrito','telefono','hospital','serologia','tipo_paciente','motivo_liberacion','ref_recibida','ref_enviada','descripCas','frecuencia','turno','distrito_cas','observacion'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            
            cadena_sin_escape = json.loads(json_data)
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)

@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_reporte_resumen_produccion(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)
            # Verificar si los valores son None o no están presentes
            fecha_inicial = objeto_python.get("fecha_inicial")
            fecha_final = objeto_python.get("fecha_final")
            # Importar las clases necesarias
            cursor = connection.cursor()

            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_reporte_resumen_produccion(%s,%s)'
            cursor.execute(sql, [fecha_inicial,fecha_final])

            # Obtener todos los resultados
            resultados = cursor.fetchall()

            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('descripCas', 'paciente','tipo_paciente','nro_sesiones'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)


@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_lista_capacidad_cupos(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)
            # Verificar si los valores son None o no están presentes
            cas_id = objeto_python.get("cas_id")
            # Importar las clases necesarias
            cursor = connection.cursor()

            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_lista_capacidad_cupos(%s)'
            cursor.execute(sql, [cas_id])

            # Obtener todos los resultados
            resultados = cursor.fetchall()

            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('turno', 'frecuencia','serologia','numeroPuesto','paciente'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)


@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_actualizacion_capacidad_cupos(request):
    if request.method == 'POST':
        try:
             # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)
            # Verificar si los valores son None o no están presentes
            cas_id = objeto_python.get("cas_id")
            turno_1 = objeto_python.get("turno1")
            frecuencia_1 = objeto_python.get("frecuencia1")
            serologia_1 = objeto_python.get("serologia1")
            numeroPuesto = objeto_python.get("numeroPuesto")
            frecuencia_2 = objeto_python.get("frecuencia")
            serologia_2 = objeto_python.get("serologia")
            usuario_reg = objeto_python.get("usuario_reg")
            # Importar las clases necesarias
            cursor = connection.cursor()
            # Importar las clases necesarias y definir la consulta SQL
            with connection.cursor() as cursor:
                sql = 'CALL public.generar_actualizacion_capacidad_cupos(%s,%s,%s,%s,%s,%s,%s,%s)'
                cursor.execute(sql,[cas_id,turno_1,frecuencia_1,serologia_1,numeroPuesto,frecuencia_2,serologia_2,usuario_reg])

            # Devolver la respuesta JSON
            return JsonResponse("Exito", safe=False)

        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)


@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_validacion_cupo(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)
            # Verificar si los valores son None o no están presentes
            clinica = objeto_python.get("cas_id")
            turno_1 = objeto_python.get("turno")
            frecuencia_1 = objeto_python.get("frecuencia")
            serologia_1 = objeto_python.get("serologia")
            usuario = objeto_python.get("usuario_reg")
            # Importar las clases necesarias
            cursor = connection.cursor()

            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_validacion_cupo(%s,%s,%s,%s,%s)'
            cursor.execute(sql, [clinica,turno_1,frecuencia_1,serologia_1,usuario])

            # Obtener todos los resultados
            resultados = cursor.fetchall()

            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('turno'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)


@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_reporte_asginacion_cupos(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)
            # Verificar si los valores son None o no están presentes
            clinica = objeto_python.get("clinica_id")
            # Importar las clases necesarias
            cursor = connection.cursor()

            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_reporte_asginacion_cupos(%s)'
            cursor.execute(sql, [clinica])

            # Obtener todos los resultados
            resultados = cursor.fetchall()

            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('paciente','num_doc','fecha_reg','usuario_reg','estado','turno','frecuencia','puesto'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)

@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_lista_protocolo_anemia(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)
            # Verificar si los valores son None o no están presentes
            fecha = objeto_python.get("fecha")
            # Importar las clases necesarias
            cursor = connection.cursor()

            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_lista_protocolo_anemia(%s)'
            cursor.execute(sql, [fecha])

            # Obtener todos los resultados
            resultados = cursor.fetchall()

            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('paciente','num_doc','hemoglobina','dosis_epo','via','ferretina','saturacion_trans','dosis_hierro','observaciones','id','sala','autogenerado','dosis_epo_ban', 'via_ban' , 'dosis_hierro_ban'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)

@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_lista_protocolo_tmo(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)
            # Verificar si los valores son None o no están presentes
            fecha = objeto_python.get("fecha")
            # Importar las clases necesarias
            cursor = connection.cursor()

            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_lista_protocolo_tmo(%s)'
            cursor.execute(sql, [fecha])

            # Obtener todos los resultados
            resultados = cursor.fetchall()

            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('PACIENTE',
                'NUM_DOC',
                'CALCIO_PRE',
                'CALCIO_POST',
                'FOSFORO_PRE',
                'FOSFORO_POST',
                'PARATOHORMONA',
                'CARBONATO_CALCIO',
                'SEVELAMERO',
                'CALCITRIOL_EV',
                'CALCITRIOL_VO',
                'PARICALCITOL',
                'CINACALCET',
                'CALCIO_ENDIALIZADO',
                'OBSERVACIONES','ID',
                'SALA','AUTOGENERADO','CARBONATO_CALCIO_BAN'
                ,'SEVELAMERO_BAN'
                ,'CALCITRIOL_EV_BAN'
                ,'CALCITRIOL_VO_BAN'
                ,'PARICALCITOL_BAN'
                ,'CINACALCET_BAN'
                ,'CALCIO_ENDIALIZADO_BAN'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)

@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_lista_protocolo_nutricion(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)
            # Verificar si los valores son None o no están presentes
            fecha = objeto_python.get("fecha")
            # Importar las clases necesarias
            cursor = connection.cursor()

            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_lista_protocolo_nutricion(%s)'
            cursor.execute(sql, [fecha])

            # Obtener todos los resultados
            resultados = cursor.fetchall()

            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('PACIENTE',
                'NUM_DOC',
                'ALBUMINA',
                'MASA_CORPORAL',
                'MASA_MUSCULAR',
                'INGESTA_PROT_CAL',
                'BIOQUIMICA',
                'TOTAL_CRITERIOS',
                'DIAGNOSTICO',
                'SNO',
                'ADMINISTRACION',
                'ID',
                'OBSERVACIONES',
                'SALA',
                'AUTOGENERADO','MASA_CORPORAL_BAN',
                'MASA_MUSCULAR_BAN',
                'INGESTA_PROT_CAL_BAN',
                'BIOQUIMICA_BAN',
                'TOTAL_CRITERIOS_BAN',
                'DIAGNOSTICO_BAN',
                'SNO_BAN',
                'ADMINISTRACION_BAN'), fila)))
            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)


@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_lista_protocolo_hta(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)
            # Verificar si los valores son None o no están presentes
            fecha = objeto_python.get("fecha")
            # Importar las clases necesarias
            cursor = connection.cursor()

            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_lista_protocolo_hta(%s)'
            cursor.execute(sql, [fecha])

            # Obtener todos los resultados
            resultados = cursor.fetchall()

            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('PACIENTE',
                'NUM_DOC',
                'PAS',
                'PAD',
                'IECA',
                'DOSIS_IECA',
                'ARA_2',
                'DOSIS_ARA_2',
                'CALCIO_ANTAGON',
                'DOSIS_CALCIO_ANTAGON',
                'ACC_SNC',
                'DOSIS_ACC_SNC',
                'BETABLOQ',
                'DOSIS_BETABLOQ',
                'ID',
                'SALA',
                'AUTOGENERADO','OBSERVACIONES'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)

@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_actualizar_protocolo_anemia(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)
            # Verificar si los valores son None o no están presentes
            id = objeto_python.get("id")
            dosis_epo = objeto_python.get("dosis_epo")
            via = objeto_python.get("via")
            dosis_hierro = objeto_python.get("dosis_hierro")
            observaciones = objeto_python.get("observaciones")
            dosis_epo_ban = objeto_python.get("dosis_epo_ban")
            via_ban = objeto_python.get("via_ban")
            dosis_hierro_ban = objeto_python.get("dosis_hierro_ban")

            # Importar las clases necesarias
            cursor = connection.cursor()

            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_actualizar_protocolo_anemia(%s,%s,%s,%s,%s,%s,%s,%s)'
            cursor.execute(sql, [id,dosis_epo,via,dosis_hierro,observaciones,dosis_epo_ban,via_ban,dosis_hierro_ban])

            # Obtener todos los resultados
            resultados = cursor.fetchall()

            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('paciente'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)


@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_actualizar_protocolo_tmo(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)
            # Verificar si los valores son None o no están presentes

            id = objeto_python.get("id")
            CARBONATO_CALCIO = objeto_python.get("CARBONATO_CALCIO")
            SEVELAMERO = objeto_python.get("SEVELAMERO")
            CALCITRIOL_EV = objeto_python.get("CALCITRIOL_EV")
            CALCITRIOL_VO = objeto_python.get("CALCITRIOL_VO")
            PARICALCITOL = objeto_python.get("PARICALCITOL")
            CINACALCET = objeto_python.get("CINACALCET")
            CALCIO_ENDIALIZADO = objeto_python.get("CALCIO_ENDIALIZADO")
            OBSERVACIONES = objeto_python.get("OBSERVACIONES")
            CARBONATO_CALCIO_BAN = objeto_python.get("CARBONATO_CALCIO_BAN")
            SEVELAMERO_BAN = objeto_python.get("SEVELAMERO_BAN")
            CALCITRIOL_EV_BAN = objeto_python.get("CALCITRIOL_EV_BAN")
            CALCITRIOL_VO_BAN = objeto_python.get("CALCITRIOL_VO_BAN")
            PARICALCITOL_BAN = objeto_python.get("PARICALCITOL_BAN")
            CINACALCET_BAN = objeto_python.get("CINACALCET_BAN")
            CALCIO_ENDIALIZADO_BAN = objeto_python.get("CALCIO_ENDIALIZADO_BAN")

            # Importar las clases necesarias
            cursor = connection.cursor()

            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_actualizar_protocolo_tmo(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)'
            cursor.execute(sql, [id,CARBONATO_CALCIO,SEVELAMERO,CALCITRIOL_EV,CALCITRIOL_VO,PARICALCITOL,CINACALCET,CALCIO_ENDIALIZADO,OBSERVACIONES,CARBONATO_CALCIO_BAN,SEVELAMERO_BAN,CALCITRIOL_EV_BAN,CALCITRIOL_VO_BAN,PARICALCITOL_BAN,CINACALCET_BAN,CALCIO_ENDIALIZADO_BAN])

            # Obtener todos los resultados
            resultados = cursor.fetchall()

            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('paciente'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)

@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_actualizar_protocolo_nutricion(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)
            # Verificar si los valores son None o no están presentes
            ID = objeto_python.get("ID")
            MASA_CORPORAL = objeto_python.get("MASA_CORPORAL")
            MASA_MUSCULAR = objeto_python.get("MASA_MUSCULAR")
            INGESTA_PROT_CAL = objeto_python.get("INGESTA_PROT_CAL")
            BIOQUIMICA = objeto_python.get("BIOQUIMICA")
            TOTAL_CRITERIOS = objeto_python.get("TOTAL_CRITERIOS")
            DIAGNOSTICO = objeto_python.get("DIAGNOSTICO")
            SNO = objeto_python.get("SNO")
            ADMINISTRACION = objeto_python.get("ADMINISTRACION")
            OBSERVACIONES = objeto_python.get("OBSERVACIONES")
            MASA_CORPORAL_BAN = objeto_python.get("MASA_CORPORAL_BAN")
            MASA_MUSCULAR_BAN = objeto_python.get("MASA_MUSCULAR_BAN")
            INGESTA_PROT_CAL_BAN = objeto_python.get("INGESTA_PROT_CAL_BAN")
            BIOQUIMICA_BAN = objeto_python.get("BIOQUIMICA_BAN")
            TOTAL_CRITERIOS_BAN = objeto_python.get("TOTAL_CRITERIOS_BAN")
            DIAGNOSTICO_BAN = objeto_python.get("DIAGNOSTICO_BAN")
            SNO_BAN = objeto_python.get("SNO_BAN")
            ADMINISTRACION_BAN = objeto_python.get("ADMINISTRACION_BAN")


            # Importar las clases necesarias
            cursor = connection.cursor()

            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_actualizar_protocolo_nutricion(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)'
            cursor.execute(sql, [ID,MASA_CORPORAL,MASA_MUSCULAR,INGESTA_PROT_CAL,BIOQUIMICA,TOTAL_CRITERIOS,DIAGNOSTICO,SNO,ADMINISTRACION,OBSERVACIONES,MASA_CORPORAL_BAN,MASA_MUSCULAR_BAN,INGESTA_PROT_CAL_BAN,BIOQUIMICA_BAN,TOTAL_CRITERIOS_BAN,DIAGNOSTICO_BAN,SNO_BAN,ADMINISTRACION_BAN])

            # Obtener todos los resultados
            resultados = cursor.fetchall()

            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('paciente'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)



@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_actualizar_protocolo_hta(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)
            
            # Verificar si los valores son None o no están presentes
            ID = objeto_python.get("ID")
            PAS = objeto_python.get("PAS")
            PAD = objeto_python.get("PAD")
            IECA = objeto_python.get("IECA")
            DOSIS_IECA = objeto_python.get("DOSIS_IECA")
            ARA_2 = objeto_python.get("ARA_2")
            DOSIS_ARA_2 = objeto_python.get("DOSIS_ARA_2")
            CALCIO_ANTAGON = objeto_python.get("CALCIO_ANTAGON")
            DOSIS_CALCIO_ANTAGON = objeto_python.get("DOSIS_CALCIO_ANTAGON")
            ACC_SNC = objeto_python.get("ACC_SNC")
            DOSIS_ACC_SNC = objeto_python.get("DOSIS_ACC_SNC")
            BETABLOQ = objeto_python.get("BETABLOQ")
            DOSIS_BETABLOQ = objeto_python.get("DOSIS_BETABLOQ")
            OBSERVACIONES = objeto_python.get("OBSERVACIONES")

            # Importar las clases necesarias
            cursor = connection.cursor()

            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_actualizar_protocolo_hta(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)'
            cursor.execute(sql, [ID,PAS,PAD,IECA,DOSIS_IECA,ARA_2,DOSIS_ARA_2,CALCIO_ANTAGON,DOSIS_CALCIO_ANTAGON,ACC_SNC,DOSIS_ACC_SNC,BETABLOQ,DOSIS_BETABLOQ,OBSERVACIONES])

            # Obtener todos los resultados
            resultados = cursor.fetchall()

            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('paciente'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)



@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_lista_pacientes_laboratorio(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)

            # Importar las clases necesarias
            cursor = connection.cursor()
            periodo = objeto_python.get("periodo")

            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_lista_pacientes_laboratorio(%s)'
            cursor.execute(sql, [periodo])

            # Obtener todos los resultados
            resultados = cursor.fetchall()

            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('dni_paciente','autogenerado','paciente','sala'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)


@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_actualizar_sala_laboratorio(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)

            # Importar las clases necesarias
            cursor = connection.cursor()
            autogenerado = objeto_python.get("autogenerado")
            mes = objeto_python.get("mes")
            periodo = objeto_python.get("periodo")
            sala_correcta = objeto_python.get("sala_correcta")

            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_actualizar_sala_laboratorio(%s,%s,%s,%s)'
            cursor.execute(sql, [autogenerado,mes,periodo,sala_correcta])

            # Obtener todos los resultados
            resultados = cursor.fetchall()

            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('dni_paciente'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)


@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_reporte_laboratorio_historico(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)

            # Importar las clases necesarias
            cursor = connection.cursor()
            fecha_inicio = objeto_python.get("fecha_inicio")
            fecha_fin = objeto_python.get("fecha_fin")
            autogenerado = objeto_python.get("autogenerado")

            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_reporte_laboratorio_historico(%s,%s,%s)'
            cursor.execute(sql, [fecha_inicio,fecha_fin,autogenerado])

            # Obtener todos los resultados
            resultados = cursor.fetchall()
            
            ip_cliente = request.META.get('REMOTE_ADDR')
            
            print(f"ip: {ip_cliente}")
            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('FECHA_RESULTADO', 'NUM_DOC', 'PACIENTE', 'AUTOGENERADO', 'SALA',
                'ALBUMINA', 'CALCIO_TOTAL', 'CALCIO_POST', 'CALCIO_PRE', 'CALCIO_IONIZADO', 
                'CLORURO', 'COLESTEROL', 'CREATININA', 'CREATININA_POST', 'CREATININA_PRE',
                'FERRITINA', 'GLUCOSA', 'FIERRO', 'COLESTEROL_HDL', 'COLESTEROL_LDL',
                'PARATOHORMONA', 'FOSFATASA_ALCALINA', 'FOSFORO', 'FOSFORO_POST',
                'FOSFORO_PRE', 'POTASIO', 'SODIO', 'TRANSAMINASA_GLUTAMICO_PIRUVIC',
                'TRANSAMINASA_GLUTAMICO_OXALACE', 'TRANSFERRINA', 'TRIGLICERIDOS', 'UREA',
                'ACIDO_URICO', 'GLOBULOS_BLANCOS', 'GLOBULOS_ROJOS', 'HEMATOCRITO',
                'HEMOGLOBINA', 'PROTEINA_C_REACTIVA', 'ANTIGENO_HEPATIT', 'ANTICUERPO_HEPATITIS_CORE',
                'ANTICUERPO_HEPATITIS_C','UREA_PRE','UREA_POST','HCO3','ANTICUERPO_HEPATITIS_B',
                'ANTIGENO_SUPERFICIE_HEPAT','ANTICUERPO_EPSILON','ACIDO_FOLICO','VITAMINA_B12','PROTEINAS',
                'SEROLOGIA','COD_IPRESS','IPRESS','TGO','ABASTONADOS','BASOFILOS','CONC_HB_CORPUSCULAR_MEDIA',
                'EOSINOFILOS','HB_CORPUSCULAR_MEDIA','LINFOCITOS','MONOCITOS','NRCB','PCT','PDW','RDW_CV','RDW_SD',
                'RECUENTO_PLAQUETAS','SEGMENTADOS','VOLUMEN_CORPUSCULAR_MEDIO','VOLUMEN_PLAQUETARIO_MEDIO',
                'PARATHORMONA_INTACTA','KTV', 'ANTICUERPO_HIV','POR_SATURACION','BETA_2','VITAMINA_D','TRU'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)


@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_consulta_resultado_anemia(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)

            # Importar las clases necesarias
            cursor = connection.cursor()
            fecha = objeto_python.get("fecha")
            autogenerado = objeto_python.get("autogen")

            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_consulta_resultado_anemia(%s,%s)'
            cursor.execute(sql, [fecha,autogenerado])

            # Obtener todos los resultados
            resultados = cursor.fetchall()
            
            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('paciente','num_doc','hemoglobina','dosis_epo','via','ferretina','saturacion_trans','dosis_hierro','observaciones','id','sala','autogenerado'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)


@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_consulta_resultado_hta(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)

            # Importar las clases necesarias
            cursor = connection.cursor()
            fecha = objeto_python.get("fecha")
            autogenerado = objeto_python.get("autogen")

            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_consulta_resultado_hta(%s,%s)'
            cursor.execute(sql, [fecha,autogenerado])

            # Obtener todos los resultados
            resultados = cursor.fetchall()
            
            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('PACIENTE',
                'NUM_DOC',
                'PAS',
                'PAD',
                'IECA',
                'DOSIS_IECA',
                'ARA_2',
                'DOSIS_ARA_2',
                'CALCIO_ANTAGON',
                'DOSIS_CALCIO_ANTAGON',
                'ACC_SNC',
                'DOSIS_ACC_SNC',
                'BETABLOQ',
                'DOSIS_BETABLOQ',
                'ID',
                'SALA',
                'AUTOGENERADO','OBSERVACIONES'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)


@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_consulta_resultado_nutricion(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)

            # Importar las clases necesarias
            cursor = connection.cursor()
            fecha = objeto_python.get("fecha")
            autogenerado = objeto_python.get("autogen")

            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_consulta_resultado_nutricion(%s,%s)'
            cursor.execute(sql, [fecha,autogenerado])

            # Obtener todos los resultados
            resultados = cursor.fetchall()
            
            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('PACIENTE',
                'NUM_DOC',
                'ALBUMINA',
                'MASA_CORPORAL',
                'MASA_MUSCULAR',
                'INGESTA_PROT_CAL',
                'BIOQUIMICA',
                'TOTAL_CRITERIOS',
                'DIAGNOSTICO',
                'SNO',
                'ADMINISTRACION',
                'ID',
                'OBSERVACIONES',
                'SALA',
                'AUTOGENERADO'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)


@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_consulta_resultado_tmo(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)

            # Importar las clases necesarias
            cursor = connection.cursor()
            fecha = objeto_python.get("fecha")
            autogenerado = objeto_python.get("autogen")

            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_consulta_resultado_tmo(%s,%s)'
            cursor.execute(sql, [fecha,autogenerado])

            # Obtener todos los resultados
            resultados = cursor.fetchall()
            
            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('PACIENTE',
                'NUM_DOC',
                'CALCIO',
                'FOSFORO',
                'PARATOHORMONA',
                'CARBONATO_CALCIO',
                'SEVELAMERO',
                'CALCITRIOL_EV',
                'CALCITRIOL_VO',
                'PARICALCITOL',
                'CINACALCET',
                'CALCIO_ENDIALIZADO',
                'OBSERVACIONES','ID',
                'SALA','AUTOGENERADO'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)


@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_actualizar_censo_paciente(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)

            # Importar las clases necesarias
            cursor = connection.cursor()
            id_paciente = objeto_python.get("paciente_id")
            telefono = objeto_python.get("telefono")
            telefonoAlterno = objeto_python.get("telefonoAlterno")
            latitud = objeto_python.get("latitud")
            longitud = objeto_python.get("longitud")
            direccion = objeto_python.get("direccion")
            ubigeo_id = objeto_python.get("ubigeo_id")
            clinica = objeto_python.get("clinica")
            turno = objeto_python.get("turno")
            secuencia = objeto_python.get("secuencia")
            fecha_reg = objeto_python.get("fecha_reg")

            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_actualizar_censo_paciente(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)'
            cursor.execute(sql, [id_paciente,telefono,telefonoAlterno,latitud,longitud,direccion,ubigeo_id,clinica,turno,secuencia,fecha_reg])

            # Obtener todos los resultados
            resultados = cursor.fetchall()
            
            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('PACIENTE'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)


@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_ingresar_censo_paciente(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)
            ip = request.META.get('REMOTE_ADDR')
            #print(ip)
            # Importar las clases necesarias
            cursor = connection.cursor()
            id_paciente = objeto_python.get("paciente_id")
            telefonoAlterno = objeto_python.get("telefonoAlterno")
            latitud = objeto_python.get("latitud")
            longitud = objeto_python.get("longitud")
            direccion = objeto_python.get("direccion")
            id_ubigeo = objeto_python.get("ubigeo_id")
            dni = objeto_python.get("num_doc")
            nombres = objeto_python.get("nombres")
            ape_pat = objeto_python.get("ape_pat")
            ape_mat = objeto_python.get("ape_mat")
            telefono = objeto_python.get("telefono")
            id_clinica = objeto_python.get("clinica")
            serologia = objeto_python.get("serologia")
            frecuencia = objeto_python.get("secuencia")
            turno = objeto_python.get("turno")
            host = ip
            usuario = objeto_python.get("usuario")
            tiempoTratamiento = objeto_python.get("tiempoTratamiento")
            distanciaCentro = objeto_python.get("distanciaCentro")
            tiempoLlegada = objeto_python.get("tiempoLlegada")
            gastoTransporte = objeto_python.get("gastoTransporte")
            atencionCercana = objeto_python.get("atencionCercana")
            tipoAccesoVascular = objeto_python.get("tipoAccesoVascular")
            medicacion = objeto_python.get("medicacion")
            terapia = objeto_python.get("terapia")
            satisfaccion = objeto_python.get("satisfaccion")
            medioTransporte = objeto_python.get("medioTransporte")
            medicoTurno = objeto_python.get("medicoTurno")
            numeroMedico = objeto_python.get("numeroMedico")
            motivoFalta = objeto_python.get("motivoFalta")

            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_ingresar_censo_paciente(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)'
            cursor.execute(sql, [id_paciente,dni,nombres,ape_pat,ape_mat,telefono,telefonoAlterno,latitud,longitud,direccion,id_ubigeo,id_clinica,serologia,frecuencia,turno,host,usuario,tiempoTratamiento,distanciaCentro,tiempoLlegada,gastoTransporte,atencionCercana,tipoAccesoVascular,medicacion,terapia,satisfaccion,medioTransporte,medicoTurno,numeroMedico,motivoFalta])
            
            # Obtener todos los resultados
            resultados = cursor.fetchall()
            
            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('PACIENTE'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)


@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_crear_paciente_cupo(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)
            #print(ip)
            # Importar las clases necesarias
            cursor = connection.cursor()
            num_doc = objeto_python.get("num_doc")
            ape_pat = objeto_python.get("ape_pat")
            ape_mat = objeto_python.get("ape_mat")
            nombres = objeto_python.get("nombres")
            hospital_id = objeto_python.get("hospital_id")
            latitud = objeto_python.get("latitud")
            longitud = objeto_python.get("longitud")
            turno = objeto_python.get("turno")
            frecuencia = objeto_python.get("frecuencia")
            tipoPuesto = objeto_python.get("tipoPuesto")
            fecha_reg = objeto_python.get("fecha_reg")
            cas_id = objeto_python.get("cas_id")

            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_crear_paciente_cupo(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)'
            cursor.execute(sql, [num_doc,ape_pat,ape_mat,nombres,hospital_id,latitud,longitud,turno,frecuencia,tipoPuesto,fecha_reg,cas_id])

            # Obtener todos los resultados
            resultados = cursor.fetchall()
            
            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('PACIENTE'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)



@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_lista_usuario_censo(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)
            
            cursor = connection.cursor()
            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_lista_usuario_censo()'
            cursor.execute(sql, [])

            # Obtener todos los resultados
            resultados = cursor.fetchall()
            
            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('id_usuario','usuario'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)

@api_view(['POST'])
@permission_required([IsAuthenticated])
def long_poll_updates(request):
    body = request.body.decode('utf-8')
    objeto_python = json.loads(body)
    fecha = objeto_python.get("fecha")

    # Importar las clases necesarias
    cursor = connection.cursor()
    # Almacena la cantidad de registros totales la primera vez que se llama
    last_record_count = formularioCenso.objects.count()
    # Variable para simular una "actualización" en este ejemplo
    last_update_time = time.time()

    # Máximo tiempo en segundos para esperar antes de responder
    TIMEOUT = 45  # Puedes ajustar el tiempo de espera

    # Intervalo de tiempo para verificar actualizaciones
    CHECK_INTERVAL = 1  # Cada segundo

    # Tiempo de inicio de la espera
    start_time = time.time()

    while time.time() - start_time < TIMEOUT:
        # Verificar si hay una actualización disponible
        if data_updated_since(last_record_count):  # Reemplaza con tu lógica de actualización
            last_update_time = time.time()
            return get_new_data(fecha)  # get_new_data() obtiene los datos actualizados
        
        # Esperar antes de la siguiente verificación
        time.sleep(CHECK_INTERVAL)

    # Si no hubo cambios en el tiempo límite, responde sin cambios
    return JsonResponse({'hasUpdates': False})

def data_updated_since(last_record_count):
    current_count = formularioCenso.objects.count()
    
    if current_count != last_record_count:
        last_record_count = current_count  # Actualiza el conteo para futuras verificaciones
        return True  # Indica que ha habido una actualización
    return False  # No ha habido cambios

def get_new_data(fecha):

    registros_mapa = generar_reporte_censo_mapa(fecha)
    registros_tabla = reporte_censo_acumulado(fecha)
    registro_acumulado=generar_reporte_censo_resumen(fecha)
    return JsonResponse({'hasUpdates': True, 'registros_mapa': registros_mapa,'registros_tabla':registros_tabla,'registro_acumulado':registro_acumulado})

@api_view(['POST'])
@permission_required([IsAuthenticated])
def dashboardEncuesta(request):
    # Función para obtener los datos actualizados que serán enviados al frontend
    body = request.body.decode('utf-8')
    objeto_python = json.loads(body)
    fecha = objeto_python.get("fecha")
    # Aquí va tu lógica para obtener los datos
    registros_mapa = generar_reporte_censo_mapa(fecha)
    
    registros_tabla = reporte_censo_acumulado(fecha)
    registro_acumulado=generar_reporte_censo_resumen(fecha)

    return JsonResponse({'hasUpdates': True, 'registros_mapa': registros_mapa,'registros_tabla':registros_tabla,'registro_acumulado':registro_acumulado})

def generar_reporte_censo_resumen(fecha):
    # Función para obtener los datos actualizados que serán enviados al frontend
    try:
            
            cursor = connection.cursor()
            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_reporte_censo_resumen(%s)'
            cursor.execute(sql, [fecha])

            # Obtener todos los resultados
            resultados = cursor.fetchall()
            columnas = [col[0] for col in cursor.description]
            # Convertir los resultados a formato JSON
            datos = []

            """ for fila in resultados:
                datos.append(dict(zip(('cantidad','nro_registros','avance','fecha_registro'), fila))) """
            for fila in resultados:
                fila_dict = {}
                for key, value in zip(columnas, fila):
                    # Convertir Decimal a str para evitar problemas de serialización
                    if isinstance(value, Decimal):
                        fila_dict[key] = str(value)
                    else:
                        fila_dict[key] = value
                datos.append(fila_dict)

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            # Devolver la respuesta JSON
            return cadena_sin_escape
        
    except json.JSONDecodeError:
            return []

def reporte_censo_acumulado(fecha):
    # Función para obtener los datos actualizados que serán enviados al frontend
    try:
            
            cursor = connection.cursor()
            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_reporte_censo(%s)'
            cursor.execute(sql, [fecha])

            # Obtener todos los resultados
            resultados = cursor.fetchall()
            
            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('descripCas','usuario','frecuencia','turno','cantidad','numero_registro','cas_id','fecha_registro'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            
            # Devolver la respuesta JSON
            return cadena_sin_escape
        
    except json.JSONDecodeError:
            return []

def generar_reporte_censo_mapa(fecha):
    # Función para obtener los datos actualizados que serán enviados al frontend
    try:
            
            cursor = connection.cursor()
            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_reporte_censo_mapa(%s)'
            cursor.execute(sql, [fecha])

            # Obtener todos los resultados
            resultados = cursor.fetchall()
            
            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('latitud','longitud','usuario','dni','nombres'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            # Devolver la respuesta JSON
            return cadena_sin_escape
        
    except json.JSONDecodeError:
            return []


@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_crear_evento(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)

            # Importar las clases necesarias
            cursor = connection.cursor()
            p_titulo = objeto_python.get("titulo")
            p_descripcion = objeto_python.get("descripcion")
            p_fecha = objeto_python.get("fecha")
            p_hora_inicio = objeto_python.get("hora_inicio")
            p_hora_fin = objeto_python.get("hora_fin")
            p_area = objeto_python.get("area")
            p_creado_por = objeto_python.get("creado_por")
            p_participantes = objeto_python.get("participantes")
            p_estado = objeto_python.get("estado")
            p_color = objeto_python.get("color")
            p_hora_inicio_text = objeto_python.get("hora_inicio_text")
            p_hora_fin_text = objeto_python.get("hora_fin_text")
            p_importancia = objeto_python.get("importancia")
            p_sala = objeto_python.get("sala")
            p_correo_sol = objeto_python.get("correo_sol")
            p_correo_jef = objeto_python.get("correo_jef")
            p_user_created = objeto_python.get("user_created")
            p_participantes_id = objeto_python.get("participantes_id")

            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_crear_evento(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)'
            cursor.execute(sql, [p_titulo,p_descripcion,p_fecha,p_hora_inicio,p_hora_fin,p_area,p_creado_por,p_participantes,p_estado,p_color,p_hora_inicio_text,p_hora_fin_text,p_importancia,p_sala,p_correo_sol,p_correo_jef,p_user_created,p_participantes_id])
            
            # Obtener todos los resultados
            resultados = cursor.fetchall()

            # Convertir los resultados a formato JSON
            datos = []
            
            for fila in resultados:
                datos.append(dict(zip(('FECHA_RESULTADO'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)


@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_lista_eventos(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)

            # Importar las clases necesarias
            cursor = connection.cursor()

            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_lista_eventos()'
            cursor.execute(sql, [])

            # Obtener todos los resultados
            resultados = cursor.fetchall()

            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('id','name','details','date','area','estado','start','end','participantes','color','hora_inicio_text','hora_fin_text','importancia','sala','correo_sol','correo_jef','user_created','personal_nombre','area_id','area_nombre','participantes_id','personal_correo'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)


@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_lista_area_auditorio(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)

            # Importar las clases necesarias
            cursor = connection.cursor()

            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_lista_area_auditorio()'
            cursor.execute(sql, [])

            # Obtener todos los resultados
            resultados = cursor.fetchall()

            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('id','nombre'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)



@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_lista_correo_auditorio(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)

            # Importar las clases necesarias
            cursor = connection.cursor()

            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_lista_correo_auditorio()'
            cursor.execute(sql, [])

            # Obtener todos los resultados
            resultados = cursor.fetchall()

            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('id','correo'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)


@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_actualizar_evento(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)

            # Importar las clases necesarias
            cursor = connection.cursor()
            p_id = objeto_python.get("id")
            p_titulo = objeto_python.get("titulo")
            p_descripcion = objeto_python.get("descripcion")
            p_fecha = objeto_python.get("fecha")
            p_hora_inicio = objeto_python.get("hora_inicio")
            p_hora_fin = objeto_python.get("hora_fin")
            p_area = objeto_python.get("area")
            p_creado_por = objeto_python.get("creado_por")
            p_participantes = objeto_python.get("participantes")
            p_estado = objeto_python.get("estado")
            p_color = objeto_python.get("color")
            p_hora_inicio_text = objeto_python.get("hora_inicio_text")
            p_hora_fin_text = objeto_python.get("hora_fin_text")
            p_importancia = objeto_python.get("importancia")
            p_sala = objeto_python.get("sala")
            p_correo_sol=objeto_python.get("correo_sol")
            p_correo_jef=objeto_python.get("correo_jef")
            p_user_created = objeto_python.get("user_created")
            p_participantes_id = objeto_python.get("participantes_id")

            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_actualizar_evento(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)'
            cursor.execute(sql, [p_id,p_titulo,p_descripcion,p_fecha,p_hora_inicio,p_hora_fin,p_area,p_creado_por,p_participantes,p_estado,p_color,p_hora_inicio_text,p_hora_fin_text,p_importancia,p_sala,p_correo_sol,p_correo_jef,p_user_created,p_participantes_id])

            # Obtener todos los resultados
            resultados = cursor.fetchall()
            print(resultados)
            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('FECHA_RESULTADO'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)

@api_view(['POST'])
@permission_required([IsAuthenticated])
def generar_eliminar_evento(request):
    if request.method == 'POST':
        try:
            # Decodificar el cuerpo de la solicitud y cargarlo como objeto Python
            body = request.body.decode('utf-8')
            objeto_python = json.loads(body)

            # Importar las clases necesarias
            cursor = connection.cursor()
            p_id = objeto_python.get("id")

            # Definir la consulta SQL y ejecutarla
            sql = 'select * from generar_eliminar_evento(%s)'
            cursor.execute(sql, [p_id])

            # Obtener todos los resultados
            resultados = cursor.fetchall()

            # Convertir los resultados a formato JSON
            datos = []

            for fila in resultados:
                datos.append(dict(zip(('FECHA_RESULTADO'), fila)))

            # Convertir la lista de diccionarios a formato JSON
            json_data = json.dumps(datos)
            cadena_sin_escape = json.loads(json_data)
            
            # Devolver la respuesta JSON
            return JsonResponse(cadena_sin_escape, safe=False)
        
        except json.JSONDecodeError:
            return JsonResponse({"error": "Error al decodificar el cuerpo de la solicitud JSON."}, status=400)
        except KeyError as e:
            return JsonResponse({"error": f"Clave faltante en el cuerpo de la solicitud: {e}"}, status=400)
        except Exception as e:
            return JsonResponse({"error": f"Error interno del servidor: {e}"}, status=500)

    return JsonResponse({"error": "Solicitud no permitida."}, status=405)


@api_view(['POST'])
@permission_required([IsAuthenticated])
def enviar_correo_con_adjunto(request):
    # Crear un objeto EmailMessage
    body_unicode = request.body.decode('utf-8')
    result = json.loads(body_unicode)
    subject = result.get("asunto")
    message = result.get("mensaje")
    from_email = settings.EMAIL_HOST_USER
    recipient_list = [result.get("destinatario")]
    #adjunto = result.get("adjunto")
    msg = EmailMessage(subject, message, from_email, recipient_list)
    msg.content_subtype = 'html'
    msg.body = html_message=result.get("html")
    # Adjuntar un archivo
    # archivo = 'D:\Proyectos2023/HISAR/codigo/back/media/media/imgOriSal/docu.pdf'  # Ruta al archivo que deseas adjuntar
    # msg.attach_file(archivo)
    # Enviar el correo
    msg.send()
    return HttpResponse("Correo enviado exitosamente.")   
##########


@api_view(['POST'])
@permission_required([IsAuthenticated])
def UsersViewSet(request):
    body_unicode = request.body.decode('utf-8')
    result = json.loads(body_unicode)
    response = requests.post(config('URL'),auth = (config('USER'), config('PASSWORD')) ,  json = {
                        "codOpcion": result.get("codOpcion"),
                        "codTipDoc": result.get("codTipDoc"),
                        "numDoc": result.get("numDoc"),
                        "fecNacimiento":result.get("fecNacimiento")
                      })
    users = response.json()

    return JsonResponse(users)
##########

##########
@api_view(['POST'])
@permission_required([IsAuthenticated])
def SeguroViewSet(request):
    body_unicode = request.body.decode('utf-8')
    result = json.loads(body_unicode)
    response = requests.get(config('URLS004')+'?tpdoc='+result.get("tpdoc")+'&nrdoc='+result.get("nrdoc")+'&codcentro='+result.get("codcentro")+'&login='+config('LOGINS004')+'&user='+config('USERS004')+'&pass='+config('PASSWORD004'))
    Seguro = response.json()
    #print("algo",Seguro)
    return JsonResponse(Seguro)
##########

##########SERVICIO DE ACCESO ESSI - BACKLOCK
@api_view(['POST'])
@permission_required([IsAuthenticated])
def LoginViewSet(request):
    body_unicode = request.body.decode('utf-8')
    result = json.loads(body_unicode)
    response = requests.post(config('URLL001'),auth = (result.get('Username'), result.get('Password')))
    login = response.json()

    return JsonResponse(login)
##########

# Pacientes geolocalizados Temporalmente

class pacienteGeoTemViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows groups to be viewed or edited.
    """
    queryset = pacienteGeoTem.objects.all()
    serializer_class = pacienteGeoTemSerializer
    permission_classes = [permissions.IsAuthenticated]    
    filter_backends = [filters.SearchFilter]
    search_fields = ['=dni']


# Tablas Generales

class ubigeoViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows groups to be viewed or edited.
    """
    queryset = ubigeo.objects.all()
    serializer_class = ubigeoSerializer
    permission_classes = [permissions.IsAuthenticated]    
    filter_backends = [filters.SearchFilter]
    search_fields = ['=ubigeo_reniec']

""" class casViewSet(viewsets.ModelViewSet):

    queryset = cas.objects.all()
    serializer_class = casSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [filters.SearchFilter]
    search_fields = ['=tipoCas','^estado','=distrito'] """

class casViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows groups to be viewed or edited.
    """
    queryset = cas.objects.all()
    serializer_class = casSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [filters.SearchFilter]
    search_fields = ['=tipoCas', '^estado', '=distrito']

    def get_queryset(self):
        queryset = super().get_queryset()
        estado = self.request.query_params.get('estado')  # Obtén el parámetro de consulta

        if estado is not None:
            # Convertir el valor de estado a booleano
            estado_booleano = estado.lower() in ['true', '1', 'yes']
            queryset = queryset.filter(estado=estado_booleano)
        
        return queryset


class usuarioViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows groups to be viewed or edited.
    """
    queryset = usuario.objects.all()
    serializer_class = usuarioSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [filters.SearchFilter]
    search_fields = ['=usuario','=id']

class usuarioIndexViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows groups to be viewed or edited.
    """
    queryset = usuario.objects.all()
    serializer_class = usuarioSerializer  # Asigna la clase serializadora correspondiente
    
    permission_classes = [permissions.IsAuthenticated]    
    # filter_backends = [filters.SearchFilter]
    search_fields = ['']

class perfilViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows groups to be viewed or edited.
    """
    queryset = perfil.objects.all()
    serializer_class = perfilSerializer  # Asigna la clase serializadora correspondiente
    
    permission_classes = [permissions.IsAuthenticated]    
    # filter_backends = [filters.SearchFilter]
    search_fields = ['']

# Create your views here.
class maestroViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows groups to be viewed or edited.
    """
    queryset = maestro.objects.all()
    serializer_class = maestroSerializer
    permission_classes = [permissions.IsAuthenticated]    
    filter_backends = [filters.SearchFilter]
    search_fields = ['=id','=codMaestro']

class PacienteViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows groups to be viewed or edited.
    """
    queryset = paciente.objects.all()
    serializer_class = PacienteSerializer
    permission_classes = [permissions.IsAuthenticated]    
    filter_backends = [filters.SearchFilter]
    search_fields = ['=id','=num_doc']

class indexPacienteViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows groups to be viewed or edited.
    """
    serializer_class = PacienteSerializer
    permission_classes = [permissions.IsAuthenticated]    
    filter_backends = [filters.SearchFilter]
    search_fields = ['num_doc', 'nombres', 'ape_pat', 'ape_mat']

    def get_queryset(self):
        queryset = paciente.objects.all()

        num_doc = self.request.query_params.get('num_doc')
        nombre = self.request.query_params.get('nombre')

        # Aplicar filtros
        if num_doc:
            queryset = queryset.filter(num_doc__icontains=num_doc)
        if nombre:
            queryset = queryset.filter(nombres__icontains=nombre)
            
        queryset = queryset.order_by('-id')
        # Tomar una rebanada del queryset
        queryset = queryset[:100]

        return queryset

class instaladoresViewSet(viewsets.ModelViewSet):

    queryset = instaladores.objects.all()
    serializer_class = instaladoresSerializer
    permission_classes = [permissions.IsAuthenticated]    
    filter_backends = [filters.SearchFilter]
    search_fields = ['=nombre']

class serologiaPacienteViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows groups to be viewed or edited.
    """
    queryset = serologiaPaciente.objects.all()
    serializer_class = serologiaPacienteSerializer
    permission_classes = [permissions.IsAuthenticated]    
    filter_backends = [filters.SearchFilter]
    search_fields = ['=paciente__num_doc']
    
class ExamenViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows groups to be viewed or edited.
    """
    queryset = examen.objects.all().order_by('-id')
    serializer_class = ExamenSerializer
    permission_classes = [permissions.IsAuthenticated]    
    filter_backends = [filters.SearchFilter]
    search_fields = ['tipo_exa']

class ArchivoViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows groups to be viewed or edited.
    """
    queryset = archivo.objects.all().order_by('-id')
    serializer_class = ArchivoSerializer
    permission_classes = [permissions.IsAuthenticated]    
    filter_backends = [filters.SearchFilter]
    search_fields = ['=paciente__id']

class presAnemiaViewSet(viewsets.ModelViewSet):
    queryset = presAnemia.objects.all().order_by('-id')
    serializer_class = presAnemiaSerializer
    permission_classes = [permissions.IsAuthenticated]    
    filter_backends = [filters.SearchFilter]
    search_fields = ['=paciente__id','=usuario__cas__codCas']

class admiAnemiaViewSet(viewsets.ModelViewSet):
    queryset = admiAnemia.objects.all().order_by('-id')
    serializer_class = admiAnemiaSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [filters.SearchFilter]
    search_fields = ['=presAnemia__paciente__id','=usuario__cas__codCas']

class exclusionAnemiaViewSet(viewsets.ModelViewSet):
    queryset = exclusionAnemia.objects.all().order_by('-id')
    serializer_class = exclusionAnemiaSerializer
    permission_classes = [permissions.IsAuthenticated]    
    filter_backends = [filters.SearchFilter]
    search_fields = ['=paciente__id','=usuario__cas__codCas']

class movimientoAnemiaViewSet(viewsets.ModelViewSet):
    queryset = movimientoAnemia.objects.all().order_by('-id')
    serializer_class = movimientoAnemiaSerializer
    permission_classes = [permissions.IsAuthenticated]    
    filter_backends = [filters.SearchFilter]
    search_fields = ['=paciente__id']

class nutricionViewSet(viewsets.ModelViewSet):
    queryset = nutricion.objects.all().order_by('-id')
    serializer_class = nutricionSerializer
    permission_classes = [permissions.IsAuthenticated]    
    filter_backends = [filters.SearchFilter]
    search_fields = ['=paciente__id','=usuario__cas__codCas','=pacNuevo']

class valGlobalSubViewSet(viewsets.ModelViewSet):
    queryset = valGlobalSub.objects.all().order_by('-id')
    serializer_class = valGlobalSubSerializer
    permission_classes = [permissions.IsAuthenticated]    
    filter_backends = [filters.SearchFilter]
    search_fields = ['=paciente__id']


class bienpatViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows groups to be viewed or edited.
    """
    queryset = bienpat.objects.all()
    serializer_class = bienpatSerializer
    permission_classes = [permissions.IsAuthenticated]    
    filter_backends = [filters.SearchFilter]
    search_fields = ['codEti']

class dependenciaViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows groups to be viewed or edited.
    """
    queryset = dependencia.objects.all()
    serializer_class = dependenciaSerializer
    permission_classes = [permissions.IsAuthenticated]    
    filter_backends = [filters.SearchFilter]
    search_fields = ['codDep']

class ambienteViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows groups to be viewed or edited.
    """
    queryset = ambiente.objects.all()
    serializer_class = ambienteSerializer
    permission_classes = [permissions.IsAuthenticated]    
    filter_backends = [filters.SearchFilter]
    search_fields = ['=dependencia__id']

class personalViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows groups to be viewed or edited.
    """
    queryset = personal.objects.all()
    serializer_class = personalSerializer
    permission_classes = [permissions.IsAuthenticated]    
    filter_backends = [filters.SearchFilter]
    search_fields = ['=dniPer']

class bienImagViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows groups to be viewed or edited.
    """
    queryset = bienImag.objects.all()
    serializer_class = bienImagSerializer
    permission_classes = [permissions.IsAuthenticated]    
    filter_backends = [filters.SearchFilter]
    search_fields = ['=bienpat__id']

class bienPersonalViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows groups to be viewed or edited.
    """
    queryset = bienPersonal.objects.all()
    serializer_class = bienPersonalSerializer
    permission_classes = [permissions.IsAuthenticated]    
    filter_backends = [filters.SearchFilter]
    search_fields = ['=bienpat__id']

class bienAmbienteViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows groups to be viewed or edited.
    """
    queryset = bienAmbiente.objects.all()
    serializer_class = bienAmbienteSerializer
    permission_classes = [permissions.IsAuthenticated]    
    filter_backends = [filters.SearchFilter]
    search_fields = ['=bienpat__id']

class bienHadwareViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows groups to be viewed or edited.
    """
    queryset = bienHadware.objects.all()
    serializer_class = bienHadwareSerializer
    permission_classes = [permissions.IsAuthenticated]    
    filter_backends = [filters.SearchFilter]
    search_fields = ['=id']

class bienSoftwareViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows groups to be viewed or edited.
    """
    queryset = bienSoftware.objects.all()
    serializer_class = bienSoftwareSerializer
    permission_classes = [permissions.IsAuthenticated]    
    filter_backends = [filters.SearchFilter]
    search_fields = ['=id']

class bienDetalleMonitorViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows groups to be viewed or edited.
    """
    queryset = bienDetalleMonitor.objects.all()
    serializer_class = bienDetalleMonitorSerializer
    permission_classes = [permissions.IsAuthenticated]    
    filter_backends = [filters.SearchFilter]
    search_fields = ['=id']

class proveedorViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows groups to be viewed or edited.
    """
    queryset = proveedor.objects.all()
    serializer_class = proveedorSerializer
    permission_classes = [permissions.IsAuthenticated]  
    filter_backends = [filters.SearchFilter]
    search_fields = ['=rucProveedor']

class provMaqViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows groups to be viewed or edited.
    """
    queryset = provMaq.objects.all()
    serializer_class = provMaqSerializer
    permission_classes = [permissions.IsAuthenticated]    
    filter_backends = [filters.SearchFilter]
    search_fields = ['id']

class incidenciaDsiViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows groups to be viewed or edited.
    """
    queryset = incidenciaDsi.objects.all().order_by('-id')
    serializer_class = incidenciaDsiSerializer
    permission_classes = [permissions.IsAuthenticated]    
    filter_backends = [filters.SearchFilter]
    search_fields = ['=personal__dniPer','=estado__descripMaestro']

class personalVpnViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows groups to be viewed or edited.
    """
    queryset = personalVpn.objects.all().order_by('-id')
    serializer_class = personalVpnSerializer
    permission_classes = [permissions.IsAuthenticated]    
    filter_backends = [filters.SearchFilter]
    search_fields = ['=personal__dniPer']

class personalCertificadoViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows groups to be viewed or edited.
    """
    queryset = personalCertificado.objects.all().order_by('-id')
    serializer_class = personalCertificadoSerializer
    permission_classes = [permissions.IsAuthenticated]    
    filter_backends = [filters.SearchFilter]
    search_fields = ['=personal__dniPer']

class delegacionBienesEstraViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows groups to be viewed or edited.
    """
    queryset = delegacionBienesEstra.objects.all().order_by('-id')
    serializer_class = delegacionBienesEstraSerializer
    permission_classes = [permissions.IsAuthenticated]    
    filter_backends = [filters.SearchFilter]
    search_fields = ['=id']

class maestroMatSapViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows groups to be viewed or edited.
    """
    queryset = maestroMatSap.objects.all().order_by('-id')
    serializer_class = maestroMatSapSerializer
    permission_classes = [permissions.IsAuthenticated]    
    filter_backends = [filters.SearchFilter]
    search_fields = ['=codSap']

class parNuticionViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows groups to be viewed or edited.
    """
    queryset = parNuticion.objects.all().order_by('-id')
    serializer_class = parNuticionSerializer
    permission_classes = [permissions.IsAuthenticated]    
    filter_backends = [filters.SearchFilter]
    #filter_fields = ['=edad','=sexo']
    search_fields = ['=edad','^sexo']

class listaEsperaViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows groups to be viewed or edited.
    """
    queryset = listaEspera.objects.all().order_by('-id')
    serializer_class = listaEsperaSerializer
    permission_classes = [permissions.IsAuthenticated]    
    filter_backends = [filters.SearchFilter]
    search_fields = ['=distrito','=estado']

class docuContratadosViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows groups to be viewed or edited.
    """
    queryset = docuContratados.objects.all().order_by('-id')
    serializer_class = docuContratadosSerializer
    permission_classes = [permissions.IsAuthenticated]    
    filter_backends = [filters.SearchFilter]
    search_fields = ['=cas__id','^estado']
class parameCentroViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows groups to be viewed or edited.
    """
    queryset = parameCentro.objects.all().order_by('-id')
    serializer_class = parameCentroSerializer
    permission_classes = [permissions.IsAuthenticated]    
    filter_backends = [filters.SearchFilter]
    search_fields = ['=cas__id','^estado']

class parameCentroPuestoViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows groups to be viewed or edited.
    """
    queryset = parameCentroPuesto.objects.all().order_by('-id')
    serializer_class = parameCentroPuestoSerializer
    permission_classes = [permissions.IsAuthenticated]    
    filter_backends = [filters.SearchFilter]
    search_fields = ['=cas__id','^estado']

class asigCuposPacViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows groups to be viewed or edited.
    """
    queryset = asigCuposPac.objects.all().order_by('-id')
    serializer_class = asigCuposPacSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [filters.SearchFilter]
    search_fields = ['=parameCentroPuesto__cas__id','=paciente__num_doc','^estado']

class asigCuposPacIndexViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows groups to be viewed or edited.
    """
    serializer_class = asigCuposPacSerializer
    permission_classes = [permissions.IsAuthenticated]    
    filter_backends = [filters.SearchFilter]
    search_fields = ['=parameCentroPuesto__cas_id','=paciente__num_doc','^estado']

def get_queryset(self):
    queryset = asigCuposPac.objects.all()

    cas = self.request.query_params.get('cas_id')
    nombres = self.request.query_params.get('nombres')
    num_doc = self.request.query_params.get('num_doc')
    # Aplicar filtros
    if cas:
        queryset = queryset.filter(parameCentroPuesto__cas__id=cas)
    if nombres:
        queryset = queryset.filter(paciente__nombres__icontains=nombres)
    if num_doc:
        queryset = queryset.filter(paciente__num_doc__icontains=num_doc)

    # Tomar una rebanada del queryset
    queryset = queryset[:100]

    return queryset

class registroCargaViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows groups to be viewed or edited.
    """
    queryset = registroCarga.objects.all().order_by('-id')
    serializer_class = registroCargaSerializer
    permission_classes = [permissions.IsAuthenticated]    
    filter_backends = [filters.SearchFilter]
    search_fields = ['=id']
class asisPacDiarioViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows groups to be viewed or edited.
    """
    queryset = asisPacDiario.objects.all().order_by('-id')
    serializer_class = asisPacDiarioSerializer
    permission_classes = [permissions.IsAuthenticated]    
    filter_backends = [filters.SearchFilter]
    search_fields = ['=asigCuposPac__parameCentroPuesto__cas__id','=validacionAsistencia']

class asisPacDiarioAdicionalViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows groups to be viewed or edited.
    """
    queryset = asisPacDiarioAdicional.objects.all().order_by('-id')
    serializer_class = asisPacDiarioAdicionalSerializer
    permission_classes = [permissions.IsAuthenticated]    
    filter_backends = [filters.SearchFilter]
    search_fields = ['=estadoAsistencia','^cas__id','^fecha_reg']


class baseDatosProduccionViewSet(viewsets.ModelViewSet):
    queryset = baseDatosProduccion.objects.all().order_by('-id')
    serializer_class = baseDatosProduccionSerializer
    permission_classes = [permissions.IsAuthenticated]    
    filter_backends = [filters.SearchFilter]
    search_fields = ['=id',]

#BACK PARA APLICATIVOS MOVIL

class loginAppHisarViewSet(viewsets.ModelViewSet):
    queryset = loginAppHisar.objects.all().order_by('-id')
    serializer_class = loginAppHisarSerializer
    permission_classes = [permissions.IsAuthenticated]    
    filter_backends = [filters.SearchFilter]
    search_fields = ['=id','=paciente__num_doc']

class dpDiarioViewSet(viewsets.ModelViewSet):
    queryset = dpDiario.objects.all().order_by('-id')
    serializer_class = dpDiarioSerializer
    pagination_class = PageNumberPagination
    permission_classes = [permissions.IsAuthenticated]    
    filter_backends = [filters.SearchFilter]
    search_fields = ['=id','=paciente__num_doc']

class examenLaboratorioViewSet(viewsets.ModelViewSet):
    queryset = examenLaboratorio.objects.all().order_by('-id')
    serializer_class = examenLaboratorioSerializer
    permission_classes = [permissions.IsAuthenticated]    
    filter_backends = [filters.SearchFilter]
    search_fields = ['=id',]

class pacienteLocalizacionViewSet(viewsets.ModelViewSet):
    queryset = pacienteLocalizacion.objects.all().order_by('-id')
    serializer_class = pacienteLocalizacionSerializer
    permission_classes = [permissions.IsAuthenticated]    
    filter_backends = [filters.SearchFilter]
    search_fields = ['paciente__num_doc','=id','=userReg',]

class movimientoPacienteViewSet(viewsets.ModelViewSet):
    queryset = movimientoPaciente.objects.all().order_by('-id')
    serializer_class = movimientoPacienteSerializer
    permission_classes = [permissions.IsAuthenticated]    
    filter_backends = [filters.SearchFilter]
    search_fields = ['=id','=cas__id']

class solicitudViewSet(viewsets.ModelViewSet):
    queryset = solicitud.objects.all().order_by('-id')
    serializer_class = solicitudSerializer
    permission_classes = [permissions.IsAuthenticated]    
    filter_backends = [filters.SearchFilter]
    search_fields = ['=paciente__num_doc','=respuesta','^cas__id']

class correosViewSet(viewsets.ModelViewSet):
    queryset = correos.objects.all().order_by('-id')
    serializer_class = correosSerializer
    permission_classes = [permissions.IsAuthenticated]    
    filter_backends = [filters.SearchFilter]
    search_fields = ['=id',]

class unidadAtencionViewSet(viewsets.ModelViewSet):
    queryset = unidadAtencion.objects.all().order_by('-id')
    serializer_class = unidadAtencionSerializer
    permission_classes = [permissions.IsAuthenticated]    
    filter_backends = [filters.SearchFilter]
    search_fields = ['=id',]

class programacionTurnoViewSet(viewsets.ModelViewSet):
    queryset = programacionTurno.objects.all().order_by('-id')
    serializer_class = programacionTurnoSerializer
    permission_classes = [permissions.IsAuthenticated]    
    filter_backends = [filters.SearchFilter]
    search_fields = ['=id',]


class incidenciaEnfermeriaCabeceraViewSet(viewsets.ModelViewSet):
    queryset = incidenciaEnfermeriaCabecera.objects.all().order_by('-id')
    serializer_class = incidenciaEnfermeriaCabeceraSerializer
    permission_classes = [permissions.IsAuthenticated]    
    filter_backends = [filters.SearchFilter]
    search_fields = ['=id',]

class procedimientoEnfermeriaViewSet(viewsets.ModelViewSet):
    queryset = procedimientoEnfermeria.objects.all().order_by('-id')
    serializer_class = procedimientoEnfermeriaSerializer
    permission_classes = [permissions.IsAuthenticated]    
    filter_backends = [filters.SearchFilter]
    search_fields = ['=id',]

class incidenciaEnfermeriaDetalleViewSet(viewsets.ModelViewSet):
    queryset = incidenciaEnfermeriaDetalle.objects.all().order_by('-id')
    serializer_class = incidenciaEnfermeriaDetalleSerializer
    permission_classes = [permissions.IsAuthenticated]    
    filter_backends = [filters.SearchFilter]
    search_fields = ['=id',]



class formularioCambioClinicaViewSet(viewsets.ModelViewSet):
    queryset = formularioCambioClinica.objects.all()
    serializer_class = formularioCambioClinicaSerializer
    permission_classes = [permissions.IsAuthenticated] 
    filter_backends = [filters.SearchFilter]
    search_fields = ['=estado']
    def get_queryset(self):
        queryset = super().get_queryset()
        search_param = self.request.query_params.get('search', None)
        if search_param == "":
            queryset = queryset.filter(estado__isnull=True)
        return queryset

class formularioCapacitacionViewSet(viewsets.ModelViewSet):
    queryset = formularioCapacitacion.objects.all()
    serializer_class = formularioCapacitacionSerializer
    permission_classes = [permissions.IsAuthenticated] 
    filter_backends = [filters.SearchFilter]
    search_fields = ['=certificado']


class hospitalViewSet(viewsets.ModelViewSet):
    queryset = hospital.objects.all()
    serializer_class = hospitalSerializer
    permission_classes = [permissions.IsAuthenticated] 
    filter_backends = [filters.SearchFilter]
    search_fields = ['=estado']
    
class medicoViewSet(viewsets.ModelViewSet):
    queryset = medico.objects.all()
    serializer_class = medicoSerializer
    permission_classes = [permissions.IsAuthenticated] 
    filter_backends = [filters.SearchFilter]
    search_fields = ['=estado']

class laboratorioViewSet(viewsets.ModelViewSet):
    queryset = laboratorio.objects.all()
    serializer_class = laboratorioSerializer
    permission_classes = [permissions.IsAuthenticated] 
    filter_backends = [filters.SearchFilter]
    search_fields = ['=estado']


class protocoloAnemiaViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows groups to be viewed or edited.
    """
    serializer_class = protocoloAnemiaSerializer
    permission_classes = [permissions.IsAuthenticated]    
    filter_backends = [filters.SearchFilter]
    search_fields = ['=fecha_resultado']

    def get_queryset(self):
        queryset = protocoloAnemia.objects.all()

        fecha_resultado = self.request.query_params.get('fecha_resultado')
        
        # Aplicar filtros
        if fecha_resultado:
            queryset = queryset.filter(fecha_resultado__startswith=fecha_resultado)

        return queryset

class protocoloTmoViewSet(viewsets.ModelViewSet):
    queryset = protocoloTmo.objects.all()
    serializer_class = protocoloTmoSerializer
    permission_classes = [permissions.IsAuthenticated] 
    filter_backends = [filters.SearchFilter]
    search_fields = ['=fecha_resultado']

class protocoloNutricionViewSet(viewsets.ModelViewSet):
    queryset = protocoloNutricion.objects.all()
    serializer_class = protocoloNutricionSerializer
    permission_classes = [permissions.IsAuthenticated] 
    filter_backends = [filters.SearchFilter]
    search_fields = ['=fecha_resultado']


def carga_ktv_archivo(request):
    if request.method == 'POST' and 'files' in request.FILES:
        # Acceder a los archivos enviados
        archivos = request.FILES.getlist('files')  # Obtiene la lista de archivos

        try:
            periodo = None
            for file in archivos:
                file_extension = file.name.split('.')[-1].lower()

                # Variables para almacenar el contenido del archivo
                sheet = None
                fecha_celda = None

                # Verificar la extensión del archivo y cargarlo en consecuencia
                if file_extension == 'xlsx':
                    workbook = openpyxl.load_workbook(file, read_only=True, data_only=True)
                    sheet = workbook.active
                    fecha_celda = sheet['J5'].value  # Cambia 'J5' por la celda donde está la fecha

                elif file_extension == 'xls':
                    workbook = xlrd.open_workbook(file_contents=file.read())
                    sheet = workbook.sheet_by_index(0)
                    fecha_celda = sheet.cell_value(rowx=4, colx=9)  # Cambia 4,9 para corresponder a 'J5'

                # Procesar la fecha
                if isinstance(fecha_celda, str):
                    fecha_obj = datetime.strptime(fecha_celda, "%d/%m/%Y")
                    periodo = fecha_obj.strftime("%Y-%m")
                elif isinstance(fecha_celda, datetime):
                    periodo = fecha_celda.strftime("%Y-%m")
                else:
                    periodo = None

                # Iterar sobre las filas del archivo y procesarlas
                if file_extension == 'xlsx':
                    for row_idx in range(7, sheet.max_row):  # Para openpyxl
                        row = [cell for cell in sheet.iter_rows(min_row=row_idx + 1, max_row=row_idx + 1, values_only=True)][0]
                        # Asignar valores de la fila a las variables correspondientes
                        (order, num_doc, num_dni, nombre, turno, edad, talla, tipoacc, qb, urea_pre, urea_post, tiempo, 
                         peso_pre, peso_post, ktv, ktve, kt, tru, bmi, tac, pcr, dializado, fecha) = row

                        # Verificar si el valor de 'ktv' no es vacío o inválido
                        if ktv == "" or ktv is None:
                            ktv = None  # O puedes usar un valor predeterminado como 0
                        if tru == "" or tru is None:
                            tru = None  # O puedes usar un valor predeterminado como 0

                        if num_dni is not None:
                            laboratorio_existente = laboratorioKtv.objects.filter(num_doc=num_dni, periodo=periodo).first()
                            if laboratorio_existente:
                                laboratorio_existente.ktv = ktv
                                laboratorio_existente.tru = tru
                                laboratorio_existente.save()
                            else:
                                laboratorio_existente_nuevo = laboratorioKtv(
                                    num_doc=num_dni,
                                    periodo=periodo,
                                    ktv=ktv,
                                    tru=tru
                                )
                                laboratorio_existente_nuevo.save()

                elif file_extension == 'xls':
                    for row_idx in range(7, sheet.nrows):  # Para xlrd
                        row = sheet.row_values(row_idx)
                        # Asignar valores de la fila a las variables correspondientes
                        (order, num_doc, num_dni, nombre, turno, edad, talla, tipoacc, qb, urea_pre, urea_post, tiempo, 
                         peso_pre, peso_post, ktv, ktve, kt, tru, bmi, tac, pcr, dializado, fecha) = row

                        # Verificar si el valor de 'ktv' no es vacío o inválido
                        if ktv == "" or ktv is None:
                            ktv = None  # O puedes usar un valor predeterminado como 0
                        if tru == "" or tru is None:
                            tru = None  # O puedes usar un valor predeterminado como 0

                        if num_dni is not None:
                            laboratorio_existente = laboratorioKtv.objects.filter(num_doc=num_dni, periodo=periodo).first()
                            if laboratorio_existente:
                                laboratorio_existente.ktv = ktv
                                laboratorio_existente.tru = tru
                                laboratorio_existente.save()
                            else:
                                laboratorio_existente_nuevo = laboratorioKtv(
                                    num_doc=num_dni,
                                    periodo=periodo,
                                    ktv=ktv,
                                    tru=tru
                                )
                                laboratorio_existente_nuevo.save()

            return JsonResponse({'message': 'Datos guardados exitosamente.'})

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    else:
        return JsonResponse({'error': 'Debes enviar al menos un archivo.'}, status=400)

